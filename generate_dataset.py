"""
generate_dataset.py
Generates a scam-message detection dataset (scam_messages_dataset.xlsx)
for Pakistani multilingual scam patterns: Urdu, Roman Urdu, English, Mixed.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import random, hashlib, datetime

# ── raw data ────────────────────────────────────────────────────────────────
# Each tuple: (message, language, category, label)
# Languages : "Urdu" | "Roman Urdu" | "English" | "Mixed"
# Categories: "Job" | "Lottery" | "Bank" | "OTP" | "SIM Block" | "Investment"
#           | "Fake Charity" | "Impersonation" | "Prize" | "Loan"
# Labels    : "Scam" | "Safe"

SCAM_MESSAGES = [
    # ──────────────── JOB SCAMS ────────────────
    ("آپ کو پاکستان کے سب سے بڑے ادارے میں نوکری کی پیشکش ہے۔ ابھی رجسٹریشن فیس 5000 روپے جمع کروائیں۔", "Urdu", "Job", "Scam"),
    ("Aap ko Saudi Arabia mein job ka offer hai. Sirf 15,000 registration fee dein aur visa process start karein.", "Roman Urdu", "Job", "Scam"),
    ("Congratulations! You have been shortlisted for a high-paying overseas job. Pay Rs. 10,000 processing fee to confirm your seat.", "English", "Job", "Scam"),
    ("Ghar bethay 50,000 per month kamayein! Sirf registration fee Rs. 3,000 bhejein aur kaam shuru karein.", "Roman Urdu", "Job", "Scam"),
    ("آپ کا انتخاب UAE میں اعلیٰ تنخواہ والی نوکری کے لیے ہو گیا ہے۔ فوری طور پر Rs. 20,000 پروسیسنگ فیس ادا کریں۔", "Urdu", "Job", "Scam"),
    ("Work from home opportunity! Earn Rs. 80,000/month with zero experience. Registration fee: Rs. 5,000 only. Limited seats!", "English", "Job", "Scam"),
    ("Bhai aap ko Dubai mein driver ki job lag sakti hai. Sirf 25,000 fee hai. Aaj hi payment karein warna seat khatam ho jayegi.", "Roman Urdu", "Job", "Scam"),
    ("محترم صاحب، آپ کو سرکاری نوکری میں بھرتی کے لیے منتخب کیا گیا ہے۔ Rs. 8,000 فیس جمع کروائیں۔", "Urdu", "Job", "Scam"),
    ("URGENT HIRING: Multinational company needs 50 employees immediately. No interview required. Pay Rs. 7,500 to secure your position.", "English", "Job", "Scam"),
    ("Online typing job available! Earn 40,000 monthly from home. Security deposit Rs. 2,000 required. Contact now.", "English", "Job", "Scam"),
    ("Aap ki CV select ho gayi hai Qatar mein. Visa fee Rs. 30,000 abhi transfer karein. Offer sirf 24 ghantay ke liye hai.", "Roman Urdu", "Job", "Scam"),
    ("نوعمر لڑکوں اور لڑکیوں کے لیے آن لائن کمائی کا سنہرا موقع۔ صرف Rs. 1,500 رجسٹریشن فیس۔", "Urdu", "Job", "Scam"),
    ("Apko part-time data entry job milti hai. Daily 2 hours kaam karein aur 60,000 kamayein. Joining fee Rs. 4,000.", "Roman Urdu", "Job", "Scam"),
    ("Guaranteed job placement in Canada! Pay Rs. 50,000 consultancy fee now. Visa approval within 30 days.", "English", "Job", "Scam"),
    ("ملازمت کی گارنٹی! صرف Rs. 6,000 میں رجسٹریشن مکمل کریں اور اگلے ہفتے کام شروع کریں۔", "Urdu", "Job", "Scam"),
    ("Easy work from home - bas mobile se 2 ghante kaam karein aur rozana 5,000 kamayein. Register karein Rs. 2,500 mein.", "Mixed", "Job", "Scam"),
    ("Overseas Employment Bureau se notification: Aap ka naam Malaysia job lottery mein nikla hai. Processing fee Rs. 15,000 jama karwayein.", "Mixed", "Job", "Scam"),
    ("Sir/Madam, your resume has been selected for Gulf countries. Pay Rs. 12,000 for medical and documentation charges immediately.", "English", "Job", "Scam"),
    ("Khas offer! Turkey mein factory worker ki 200 posts khali hain. Registration fee sirf Rs. 8,000. Aaj hi apply karein.", "Roman Urdu", "Job", "Scam"),
    ("WhatsApp par part-time job offer: rozana 3 ghante kaam, ماہانہ 70,000 روپے کمائیں۔ پہلے Rs. 3,500 جمع کروائیں۔", "Mixed", "Job", "Scam"),

    # ──────────────── LOTTERY SCAMS ────────────────
    ("Mubarak ho! Aap ne Rs. 50,00,000 ki lottery jeet li hai. Claim karne ke liye Rs. 5,000 processing fee bhejein.", "Roman Urdu", "Lottery", "Scam"),
    ("مبارک ہو! آپ نے 1 کروڑ روپے کی قرعہ اندازی جیت لی ہے۔ انعام حاصل کرنے کے لیے Rs. 10,000 بھیجیں۔", "Urdu", "Lottery", "Scam"),
    ("Congratulations! You have WON the international lottery prize of $500,000 USD! Pay $200 processing fee to claim.", "English", "Lottery", "Scam"),
    ("Aap ka number Jeeto Pakistan show mein select hua hai! Rs. 20,00,000 ka inaam. Pehle Rs. 8,000 tax pay karein.", "Roman Urdu", "Lottery", "Scam"),
    ("PTI Lucky Draw: Aap ne 10 lakh rupees jeet liye hain! Verification ke liye Rs. 5,000 transfer karein.", "Roman Urdu", "Lottery", "Scam"),
    ("آپ کا موبائل نمبر قرعہ اندازی میں منتخب ہوا ہے۔ Rs. 30,00,000 کا انعام جیتنے کے لیے ابھی رابطہ کریں۔", "Urdu", "Lottery", "Scam"),
    ("You have been randomly selected to win Rs. 25,00,000! Reply with your CNIC number and Rs. 3,000 fee to claim.", "English", "Lottery", "Scam"),
    ("Coca Cola Lucky Draw Winner! Aap ko Rs. 15,00,000 milein ge. Processing fee Rs. 7,500 abhi bhejein.", "Mixed", "Lottery", "Scam"),
    ("Nestle Pakistan promotion mein aap ka number nikla hai! Prize money Rs. 50,00,000. Claim fee Rs. 12,000.", "Mixed", "Lottery", "Scam"),
    ("Lakhpati bana ka offer! Bs Rs. 2,000 investment karein aur 10 lakh tak jeetne ka mauka payein.", "Roman Urdu", "Lottery", "Scam"),
    ("آپ نے غیر ملکی لاٹری میں $100,000 جیتے ہیں۔ کلیم کرنے کے لیے $500 ٹرانسفر کریں۔", "Urdu", "Lottery", "Scam"),
    ("Urgent: Your lottery winnings of Rs. 75,00,000 will expire in 48 hours! Pay Rs. 15,000 now to claim before deadline.", "English", "Lottery", "Scam"),
    ("Samsung Galaxy promotion - aap ne car jeet li hai! Registration fee Rs. 10,000 pay karein aur car collect karein.", "Mixed", "Lottery", "Scam"),
    ("Jazz Cash Lucky Draw! Aap ko Rs. 5,00,000 ka cash prize mila hai. Sirf Rs. 2,500 fee se claim karein.", "Mixed", "Lottery", "Scam"),
    ("Aap ki kismat khul gayi! International lucky draw mein aap ka naam aaya hai - Rs. 1 crore prize. Processing charges Rs. 25,000.", "Roman Urdu", "Lottery", "Scam"),

    # ──────────────── BANK FRAUD ────────────────
    ("Your HBL bank account has been compromised. Click this link immediately to verify: http://hbl-verify.com/login", "English", "Bank", "Scam"),
    ("آپ کا بینک اکاؤنٹ بند ہونے والا ہے۔ فوری طور پر اس لنک پر اپنی تفصیلات اپ ڈیٹ کریں: http://bank-update.pk", "Urdu", "Bank", "Scam"),
    ("Aap ka Meezan Bank account block hone wala hai. Is link par click karke verify karein: http://meezan-secure.com", "Roman Urdu", "Bank", "Scam"),
    ("ALERT: Suspicious transaction of Rs. 45,000 detected on your UBL account. Verify at http://ubl-security.com now!", "English", "Bank", "Scam"),
    ("Aap ki debit card ki details compromise ho gayi hain. Naya card activate karne ke liye yeh link follow karein.", "Roman Urdu", "Bank", "Scam"),
    ("آپ کے اکاؤنٹ سے Rs. 78,000 کی غیر مجاز ٹرانزیکشن ہوئی ہے۔ فوری تصدیق کریں۔", "Urdu", "Bank", "Scam"),
    ("Your Allied Bank credit card has exceeded its limit. Pay Rs. 15,000 minimum to avoid legal action. Link: http://abl-pay.com", "English", "Bank", "Scam"),
    ("SARKARI NOTIFICATION: Aap ke bank account mein Rs. 25,000 government scheme ke aaye hain. Is link se withdraw karein.", "Mixed", "Bank", "Scam"),
    ("Important: Your bank KYC verification is pending. Account freeze hoga 24 ghante mein. Verify now: http://kyc-bank.pk", "Mixed", "Bank", "Scam"),
    ("Aap ka bank account Pakistan State Bank ne freeze kar diya hai. Unblock karne ke liye CNIC aur password share karein.", "Roman Urdu", "Bank", "Scam"),
    ("Dear Customer, your internet banking has been temporarily suspended. Reactivate by confirming your details at http://reactivate-bank.com", "English", "Bank", "Scam"),
    ("آپ کے بینک اکاؤنٹ کی تصدیق ضروری ہے۔ اپنے CNIC نمبر اور پاس ورڈ کا اشتراک کریں ورنہ اکاؤنٹ بند کر دیا جائے گا۔", "Urdu", "Bank", "Scam"),
    ("FBR Tax Refund: Aap ko Rs. 85,000 ka tax refund mila hai. Bank details is link par enter karein: http://fbr-refund.pk", "Mixed", "Bank", "Scam"),
    ("Your MCB account requires urgent security update. Please confirm your CNIC, account number, and password.", "English", "Bank", "Scam"),
    ("Aap ke naam par fake account operate ho raha hai. Investigation ke liye apni details share karein warna police action hoga.", "Roman Urdu", "Bank", "Scam"),
    ("Bank Alfalah Alert: Rs. 1,25,000 ka loan aap ke naam par approve hua hai. Agar aap ne apply nahi kiya to yeh link click karein.", "Mixed", "Bank", "Scam"),
    ("آپ کے اکاؤنٹ میں مشکوک سرگرمی کا پتہ چلا ہے۔ فوری طور پر کسٹمر سروس کو اپنے اکاؤنٹ کی تفصیلات فراہم کریں۔", "Urdu", "Bank", "Scam"),
    ("NADRA verification required for your bank account. Apna CNIC number aur date of birth is link par enter karein.", "Mixed", "Bank", "Scam"),
    ("BENAZIR INCOME SUPPORT: Aap ke account mein Rs. 12,000 aaye hain. Link par click karke confirm karein: http://bisp-gov.com", "Mixed", "Bank", "Scam"),
    ("Your Standard Chartered account needs immediate password reset due to security breach. Click: http://scb-reset.com", "English", "Bank", "Scam"),

    # ──────────────── OTP SCAMS ────────────────
    ("Aap ko abhi ek OTP bheja gaya hai. Woh code mujhe bata dein taake aap ka account verify ho jaye.", "Roman Urdu", "OTP", "Scam"),
    ("آپ کو بھیجا گیا OTP کوڈ کسی کے ساتھ شیئر نہ کریں۔ اگر کسی نے مانگا تو یہ دھوکہ دہی ہے۔", "Urdu", "OTP", "Scam"),
    ("JazzCash team calling: Aap ka account verify karne ke liye OTP code share karein. Hum aap ko reward dein ge.", "Roman Urdu", "OTP", "Scam"),
    ("Sir, aap ke EasyPaisa account mein masla hai. Verification ke liye jo code aaya hai woh bata dein.", "Roman Urdu", "OTP", "Scam"),
    ("IMPORTANT: Your bank sent an OTP? Please share it with our verification team to complete your KYC process.", "English", "OTP", "Scam"),
    ("Aap ka JazzCash payment bhejne ke liye OTP chahiye. Jo code aaya hai woh mujhe send karein.", "Roman Urdu", "OTP", "Scam"),
    ("آپ کا ایزی پیسہ اکاؤنٹ اپ گریڈ کرنے کے لیے OTP کوڈ درکار ہے۔ براہ کرم کوڈ فراہم کریں۔", "Urdu", "OTP", "Scam"),
    ("Your Rs. 50,000 transaction is pending OTP verification. Please share the OTP you received to complete the transfer.", "English", "OTP", "Scam"),
    ("HBL Mobile se call: Sir aap ka mobile banking activate karne ke liye OTP chahiye jo aap ko SMS se aaya hai.", "Mixed", "OTP", "Scam"),
    ("Aap ka EasyPaisa account hacked ho gaya hai! Bachane ke liye abhi OTP code share karein jo SMS pe aaya.", "Roman Urdu", "OTP", "Scam"),
    ("Government verification process: OTP share karein jo aap ko mila hai warna aap ka mobile number band ho jaye ga.", "Mixed", "OTP", "Scam"),
    ("Rs. 25,000 ka cashback aap ke liye! Bas OTP code confirm karein jo abhi aap ko mila hai.", "Mixed", "OTP", "Scam"),
    ("Technical support: Aap ke phone mein virus hai. Jo verification code aaya hai woh humein bata dein taake hum fix kar sakein.", "Roman Urdu", "OTP", "Scam"),
    ("آپ کی رقم منتقلی مکمل کرنے کے لیے OTP کوڈ کی ضرورت ہے۔ SMS پر موصول ہونے والا کوڈ بھیجیں۔", "Urdu", "OTP", "Scam"),
    ("Aap ka Telenor Microfinance account verify nahi hua. OTP code share karein warna account band ho jaye ga.", "Roman Urdu", "OTP", "Scam"),

    # ──────────────── SIM BLOCK SCAMS ────────────────
    ("PTA NOTICE: Aap ka SIM card 24 ghante mein block ho jaye ga. Bachane ke liye is number par CNIC details bhejein.", "Roman Urdu", "SIM Block", "Scam"),
    ("آپ کا جاز سم کارڈ بلاک ہونے والا ہے۔ فوری طور پر اپنا CNIC نمبر اور تفصیلات بھیجیں۔", "Urdu", "SIM Block", "Scam"),
    ("Dear Telenor subscriber, your SIM will be deactivated due to non-verification. Reply with CNIC to keep it active.", "English", "SIM Block", "Scam"),
    ("Aap ka Zong number block hone wala hai kyunke aap ne biometric verification nahi karwayi. Abhi CNIC bhejein.", "Roman Urdu", "SIM Block", "Scam"),
    ("FINAL WARNING: Your Ufone SIM will be permanently blocked in 12 hours. Send CNIC copy for re-verification.", "English", "SIM Block", "Scam"),
    ("PTA ki taraf se notice: Aap ka mobile number NADRA se verify nahi hua. CNIC details share karein warna SIM block.", "Mixed", "SIM Block", "Scam"),
    ("آپ کا موبائل نمبر پاکستان ٹیلی کمیونیکیشن اتھارٹی نے بلاک کرنے کا فیصلہ کیا ہے۔ تصدیق کے لیے رابطہ کریں۔", "Urdu", "SIM Block", "Scam"),
    ("URGENT: Aap ke naam par 5 SIM cards registered hain. Jo use nahi karte unko band karne ke liye CNIC aur details bhejein.", "Roman Urdu", "SIM Block", "Scam"),
    ("Your Jazz SIM has been flagged for illegal activity. Provide CNIC and fingerprint details to avoid blocking.", "English", "SIM Block", "Scam"),
    ("Aap ka number kal tak band ho jaye ga agar aap ne re-verification nahi karwayi. Is link par CNIC upload karein.", "Roman Urdu", "SIM Block", "Scam"),
    ("PTA Alert: Unregistered SIMs will be blocked nationwide. Apni SIM verify karwane ke liye yeh form fill karein.", "Mixed", "SIM Block", "Scam"),
    ("آپ کے نام پر درج تمام سم کارڈز بلاک کیے جا رہے ہیں۔ بحال کرنے کے لیے بائیو میٹرک تصدیق کروائیں۔", "Urdu", "SIM Block", "Scam"),
    ("Your Telenor number has been selected for mandatory SIM replacement. Pay Rs. 500 processing fee to avoid disconnection.", "English", "SIM Block", "Scam"),
    ("Aap ka Zong SIM expire ho raha hai. Rs. 1,000 recharge karein is special code se warna number kisi aur ko allot ho jaye ga.", "Roman Urdu", "SIM Block", "Scam"),
    ("NOTICE: Aap ka mobile connection illegal import se related hai. PTA office mein report karein ya CNIC bhej kar clear karein.", "Mixed", "SIM Block", "Scam"),

    # ──────────────── INVESTMENT SCAMS ────────────────
    ("Crypto investment mein Rs. 10,000 lagayein aur 30 din mein Rs. 1,00,000 kamayein. Guaranteed returns!", "Mixed", "Investment", "Scam"),
    ("آپ صرف Rs. 5,000 سرمایہ کاری کریں اور ہر ماہ 200% منافع کمائیں۔ یہ موقع دوبارہ نہیں آئے گا!", "Urdu", "Investment", "Scam"),
    ("Double your money in 15 days! Invest Rs. 20,000 and get Rs. 40,000 back. 100% guaranteed scheme.", "English", "Investment", "Scam"),
    ("Bhai, ek zabardast scheme hai. Rs. 5,000 do, har hafte Rs. 3,000 milega. Main khud 2 lakh kama chuka hoon.", "Roman Urdu", "Investment", "Scam"),
    ("Forex trading se Pakistan mein beth kar dollars kamayein! Minimum investment Rs. 15,000. Monthly profit guaranteed.", "Mixed", "Investment", "Scam"),
    ("آپ کی سرمایہ کاری پر ماہانہ 50% منافع کی گارنٹی! ابھی Rs. 25,000 سے شروع کریں۔", "Urdu", "Investment", "Scam"),
    ("Binary trading platform - invest $100 and earn $1000 daily. Pakistanis already earning lakhs! Join now.", "English", "Investment", "Scam"),
    ("Mutual fund scheme: Rs. 50,000 invest karein aur har mahine Rs. 10,000 munafa payein. SECP registered hai.", "Mixed", "Investment", "Scam"),
    ("Amazon FBA course join karein Rs. 25,000 mein aur ghar bethay lakhs kamayein. Success ki guarantee!", "Mixed", "Investment", "Scam"),
    ("Ponzi... I mean Pool investment scheme! Rs. 3,000 dalein aur daily Rs. 500 return payein. 100 log join ho chuke hain.", "Roman Urdu", "Investment", "Scam"),
    ("Gold investment plan: Sona kharidein baghair sonay ke! Rs. 10,000 monthly invest karein aur 20% profit lein.", "Mixed", "Investment", "Scam"),
    ("آئی ٹی سیکٹر میں سرمایہ کاری کا سنہرا موقع۔ Rs. 50,000 لگائیں اور سالانہ 300% ریٹرن حاصل کریں۔", "Urdu", "Investment", "Scam"),
    ("Earn passive income! Real estate investment trust - invest Rs. 1,00,000 and get Rs. 20,000 monthly rental income guaranteed.", "English", "Investment", "Scam"),
    ("TikTok se paise kamao scheme! Rs. 5,000 invest karein aur daily Rs. 1,000 kamayein by watching videos.", "Mixed", "Investment", "Scam"),
    ("Binance par trading sikhayein free mein! Lekin pehle Rs. 10,000 deposit karein as security. Profit ka 50% humara.", "Mixed", "Investment", "Scam"),

    # ──────────────── FAKE CHARITY SCAMS ────────────────
    ("سیلاب متاثرین کی مدد کریں! اپنے عطیات اس نمبر پر بھیجیں۔ ہر روپیہ ضرورت مندوں تک پہنچے گا۔", "Urdu", "Fake Charity", "Scam"),
    ("Flood victims need your help! Donate to this JazzCash number. 100% goes to affected families in Sindh.", "English", "Fake Charity", "Scam"),
    ("Ramadan Mubarak! Zakat aur Sadqa is number par bhejein. Hum orphan children ki kafaalat karte hain.", "Mixed", "Fake Charity", "Scam"),
    ("Aap ki zakat qabool ho! Is account par Rs. 50,000 zakat bhejein. Hum registered charity hain (fake registration number).", "Roman Urdu", "Fake Charity", "Scam"),
    ("Kashmir earthquake relief fund - Donate generously. EasyPaisa number: 03XX-XXXXXXX. Every rupee counts!", "English", "Fake Charity", "Scam"),
    ("یتیم بچوں کی کفالت کا ثواب کمائیں۔ ماہانہ صرف Rs. 2,000 اس نمبر پر بھیجیں۔", "Urdu", "Fake Charity", "Scam"),
    ("Eid-ul-Azha Qurbani program: Rs. 15,000 bhejein aur hum aap ki taraf se qurbani karein ge. Receipt milay gi.", "Mixed", "Fake Charity", "Scam"),
    ("Bhai, main ek NGO chalata hoon. Cancer patients ke liye donation chahiye. Rs. 500 bhi bohat hai. JazzCash pe bhejein.", "Roman Urdu", "Fake Charity", "Scam"),
    ("Winter clothing drive for Thar! Rs. 3,000 donate karein aur hum kapray distribute karein ge. Tax exemption certificate milega.", "Mixed", "Fake Charity", "Scam"),
    ("فلاح انسانیت فاؤنڈیشن: غریب خاندانوں کو راشن فراہم کرنے کے لیے عطیات دیں۔ بینک اکاؤنٹ نمبر: XXXX", "Urdu", "Fake Charity", "Scam"),

    # ──────────────── IMPERSONATION SCAMS ────────────────
    ("Assalam o Alaikum, main aap ka cousin Ali hoon. Mera phone kharab ho gaya hai. Mujhe Rs. 5,000 bhej do urgently.", "Roman Urdu", "Impersonation", "Scam"),
    ("Mom here. I'm at the hospital and need Rs. 20,000 for emergency treatment. Send to this number immediately.", "English", "Impersonation", "Scam"),
    ("السلام علیکم! میں آپ کا بھائی ہوں۔ نیا نمبر ہے۔ مجھے فوری Rs. 10,000 کی ضرورت ہے۔", "Urdu", "Impersonation", "Scam"),
    ("Bhai, yeh mera naya number hai. Purana phone kho gaya. Mujhe Rs. 15,000 send karo, kal wapis karunga.", "Roman Urdu", "Impersonation", "Scam"),
    ("This is your boss Mr. Ahmed. I'm in a meeting and need an urgent favor. Transfer Rs. 50,000 to this account right now.", "English", "Impersonation", "Scam"),
    ("Beta, main Ammi bol rahi hoon. Naye number se message kar rahi hoon. Rs. 8,000 bhej do bohat urgent hai.", "Roman Urdu", "Impersonation", "Scam"),
    ("آپ کے دوست کا نیا نمبر۔ مجھے ابھی Rs. 3,000 بھیجو، بعد میں واپس کروں گا۔", "Urdu", "Impersonation", "Scam"),
    ("Aslam o Alaikum bhai jan. Main Dubai mein hoon aur mera wallet chori ho gaya. Rs. 30,000 transfer karo please.", "Roman Urdu", "Impersonation", "Scam"),
    ("This is IT department. Your company email password needs reset. Share your current password for verification.", "English", "Impersonation", "Scam"),
    ("Sir, main aap ka dost Rashid. Family emergency hai, Rs. 25,000 chahiye. Yeh mera naya number hai, kal mil ke baat karein ge.", "Roman Urdu", "Impersonation", "Scam"),
    ("Papa here - office trip extended. Send Rs. 40,000 to colleague's account for hotel booking. Will explain later.", "English", "Impersonation", "Scam"),
    ("Aap ki bhabi ka number hai. Ghar mein emergency hai aur bhai ka phone nahi lag raha. Rs. 15,000 bhej dein.", "Roman Urdu", "Impersonation", "Scam"),

    # ──────────────── LOAN SCAMS ────────────────
    ("فوری قرضہ! صرف Rs. 2,000 پروسیسنگ فیس پر Rs. 2,00,000 تک کا لون حاصل کریں۔ کوئی ضمانت نہیں۔", "Urdu", "Loan", "Scam"),
    ("Instant loan approval! Get Rs. 5,00,000 with no collateral. Processing fee: Rs. 10,000 only. Apply now!", "English", "Loan", "Scam"),
    ("Aap ko Rs. 3,00,000 ka loan pre-approved hai. Sirf Rs. 5,000 documentation charges pay karein.", "Roman Urdu", "Loan", "Scam"),
    ("Bina byaj ka qarz chahiye? Islamic microfinance scheme mein apply karein. Processing fee Rs. 3,000.", "Mixed", "Loan", "Scam"),
    ("SBP scheme: Students ke liye Rs. 10,00,000 ka education loan. Sirf Rs. 8,000 fee se apply karein.", "Mixed", "Loan", "Scam"),
    ("آپ کو بزنس لون کی پیشکش! Rs. 10,00,000 تک 0% شرح سود پر۔ Rs. 15,000 ایڈوانس فیس ادا کریں۔", "Urdu", "Loan", "Scam"),
    ("Bad credit? No problem! Get approved for Rs. 2,00,000 personal loan. Pay Rs. 7,000 advance processing fee.", "English", "Loan", "Scam"),
    ("Ghar khareedne ka khwab poora karein! Home loan Rs. 50,00,000 tak. Application fee Rs. 20,000 only.", "Mixed", "Loan", "Scam"),
    ("Aap ka loan application approve ho gaya hai! Rs. 5,00,000 ka amount. Insurance fee Rs. 12,000 pay karein release ke liye.", "Roman Urdu", "Loan", "Scam"),
    ("Emergency loan 1 ghante mein! Rs. 50,000 se Rs. 10,00,000 tak. Advance fee sirf 2%. Aaj hi apply karein.", "Roman Urdu", "Loan", "Scam"),

    # ──────────────── PRIZE / GIFT SCAMS ────────────────
    ("Aap ne iPhone 15 Pro jeet liya hai! Delivery charges Rs. 2,500 pay karein aur phone ghar bethay payein.", "Roman Urdu", "Prize", "Scam"),
    ("Congratulations! You have won a free laptop from Daraz.pk. Pay Rs. 1,500 shipping fee to receive it.", "English", "Prize", "Scam"),
    ("آپ نے قرعہ اندازی میں نئی Honda Civic جیتی ہے! Rs. 25,000 ڈلیوری چارجز ادا کریں۔", "Urdu", "Prize", "Scam"),
    ("KFC Pakistan Anniversary: Aap ko free meal combo mila hai! Rs. 200 delivery fee is link par pay karein.", "Mixed", "Prize", "Scam"),
    ("You've been selected for a FREE Samsung TV! Just pay Rs. 3,000 customs/delivery charges to claim your prize.", "English", "Prize", "Scam"),
    ("Jazz rewards program: Aap ne 10GB free data aur Rs. 5,000 balance jeeta hai. Claim fee Rs. 500.", "Mixed", "Prize", "Scam"),
    ("Alhamdulillah! Aap ne Umrah package jeet liya hai. Sirf Rs. 50,000 visa aur ticket charges pay karein.", "Roman Urdu", "Prize", "Scam"),
    ("FoodPanda loyalty reward: Aap ko Rs. 10,000 ka voucher mila hai! Rs. 300 activation fee pay karein.", "Mixed", "Prize", "Scam"),
    ("آپ نے شاپنگ سپری مقابلے میں Rs. 1,00,000 کا شاپنگ واؤچر جیتا ہے۔ Rs. 2,000 فیس ادا کریں۔", "Urdu", "Prize", "Scam"),
   ("Daraz Mega Sale Winner! Aap ko AirPods Pro free milein ge. Shipping fee Rs. 1,000 is account par bhejein.", "Mixed", "Prize", "Scam"),

    # ──────────────── ADDITIONAL JOB SCAMS ────────────────
    ("Bilkul free course join karein aur 1 mahine mein job ki guarantee! Sirf Rs. 4,000 material charges hain.", "Roman Urdu", "Job", "Scam"),
    ("Amazon virtual assistant position open! Monthly salary $2,000 USD. Pay $50 training fee to start.", "English", "Job", "Scam"),
    ("فوری بھرتی! سرکاری اسکول میں اساتذہ کی ضرورت۔ Rs. 12,000 فیس کے ساتھ اپلائی کریں۔", "Urdu", "Job", "Scam"),
    ("Aap ko UAE ki top company se call aaya tha na? Woh log abhi bhi wait kar rahe hain. Rs. 18,000 visa fee bhejein.", "Roman Urdu", "Job", "Scam"),
    ("Data entry job: Type 10 pages daily, earn Rs. 2,000 per page. Security deposit Rs. 5,000 required first.", "English", "Job", "Scam"),
    ("آپ کو قطر میں ہوٹل کی نوکری مل سکتی ہے۔ ایجنسی فیس Rs. 35,000۔ ویزا 15 دن میں۔", "Urdu", "Job", "Scam"),
    ("YouTube channel monetization service! Rs. 8,000 pay karein aur hum aap ka channel monetize karein ge guaranteed.", "Mixed", "Job", "Scam"),
    ("Multinational corporation urgently hiring remote workers. No experience needed. Pay Rs. 6,000 onboarding fee.", "English", "Job", "Scam"),
    ("Bhai, Saudi Arabia mein electrician ki 50 posts hain. Salary 3000 SAR. Registration Rs. 10,000. Limited seats!", "Roman Urdu", "Job", "Scam"),
    ("Social media manager ki job! Ghar se kaam karein. ماہانہ Rs. 45,000۔ ٹریننگ فیس Rs. 3,000۔", "Mixed", "Job", "Scam"),

    # ──────────────── ADDITIONAL LOTTERY SCAMS ────────────────
    ("Telenor reward: Aap ko 5 lakh rupees milne wale hain! Rs. 3,000 transfer karke confirm karein.", "Roman Urdu", "Lottery", "Scam"),
    ("UK International Lottery: You are a winner of £250,000! Claim fee £150 via Western Union.", "English", "Lottery", "Scam"),
    ("آپ کا موبائل نمبر سالانہ قرعہ اندازی میں Rs. 20 لاکھ کا انعام جیتا ہے۔ ٹیکس Rs. 15,000 ادا کریں۔", "Urdu", "Lottery", "Scam"),
    ("Pepsi cap mein code mila? Mubarak ho aap ne Rs. 10,00,000 jeet liye! Processing ke liye Rs. 5,000 bhejein.", "Mixed", "Lottery", "Scam"),
    ("WhatsApp forward: Send this to 10 people and win Rs. 50,000 free balance! Offer valid till midnight.", "English", "Lottery", "Scam"),
    ("Bol TV game show mein aap ka naam nikla hai! Prize Rs. 8,00,000. Claim karne ke liye call karein.", "Roman Urdu", "Lottery", "Scam"),
    ("آپ نے آن لائن قرعہ اندازی میں نئی موٹر سائیکل جیتی ہے! Rs. 5,000 ڈلیوری فیس بھیجیں۔", "Urdu", "Lottery", "Scam"),
    ("Lucky number alert! Aap ka mobile number Rs. 3,00,000 ka winner hai. Sirf CNIC aur Rs. 4,000 bhejein.", "Mixed", "Lottery", "Scam"),
    ("Eid special draw: Rs. 1 crore ka first prize! Ticket sirf Rs. 500 mein. Abhi khareedein!", "Mixed", "Lottery", "Scam"),
    ("State Bank promotional scheme mein aap ka naam aaya hai. Rs. 5,00,000 prize. Verification fee Rs. 8,000.", "Mixed", "Lottery", "Scam"),

    # ──────────────── ADDITIONAL BANK FRAUD ────────────────
    ("Your Habib Bank account has unusual login from Dubai. If not you, click http://hbl-secure.net to lock account.", "English", "Bank", "Scam"),
    ("Aap ke Meezan account se Rs. 2,00,000 ki transaction hui hai. Agar aap ne nahi ki to link click karein: http://meezan-alert.com", "Roman Urdu", "Bank", "Scam"),
    ("آپ کے ڈیبٹ کارڈ کو اپ گریڈ کرنے کی ضرورت ہے۔ کارڈ نمبر اور CVV اس لنک پر درج کریں۔", "Urdu", "Bank", "Scam"),
    ("URGENT from State Bank: Your account has been flagged. Submit CNIC and bank password at http://sbp-verify.org", "English", "Bank", "Scam"),
    ("Aap ka credit card block kar diya gaya hai. Unblock karne ke liye call karein aur apni saari details batayein.", "Roman Urdu", "Bank", "Scam"),
    ("Bank transfer pending: Rs. 5,00,000 aap ke account mein aane wale hain. Confirm karne ke liye link click karein.", "Roman Urdu", "Bank", "Scam"),
    ("آپ کے اکاؤنٹ کی کے وائی سی (KYC) تصدیق نہیں ہوئی۔ 48 گھنٹے میں اکاؤنٹ منجمد ہو جائے گا۔", "Urdu", "Bank", "Scam"),
    ("FBR Income Tax: Aap ka refund Rs. 1,50,000 ready hai. Bank account details is link par fill karein.", "Mixed", "Bank", "Scam"),
    ("Your UBL Omni account has been upgraded to premium. Verification charges Rs. 2,000. Pay now.", "English", "Bank", "Scam"),
    ("Dear customer, your cheque of Rs. 3,00,000 has been dishonoured. Resolve by visiting http://resolve-bank.com", "English", "Bank", "Scam"),

    # ──────────────── ADDITIONAL OTP SCAMS ────────────────
    ("Aap ko jo OTP aaya hai woh humein de dein. Bank walon ne kaha hai ke verification ke liye zaroori hai.", "Roman Urdu", "OTP", "Scam"),
    ("Customer support: Sir aap ka payment process karne ke liye OTP chahiye. Jo SMS pe aaya hai woh batayein.", "Roman Urdu", "OTP", "Scam"),
    ("آپ کا بینک اکاؤنٹ ہیک ہو رہا ہے! فوری طور پر OTP کوڈ شیئر کریں تاکہ ہم اکاؤنٹ محفوظ کر سکیں۔", "Urdu", "OTP", "Scam"),
    ("EasyPaisa promotion: Apna OTP code share karein aur Rs. 5,000 free balance payein. Offer limited!", "Mixed", "OTP", "Scam"),
    ("Aap ke number par Rs. 50,000 aaye hain kisi ne bheje. Claim karne ke liye OTP code confirm karein.", "Roman Urdu", "OTP", "Scam"),
    ("Your JazzCash account will be closed if OTP verification is not completed in 1 hour. Share the code now.", "English", "OTP", "Scam"),
    ("Bhai, main bank se baat kar raha hoon. Unhone kaha OTP share karo tabhi transfer hoga. Code batao.", "Roman Urdu", "OTP", "Scam"),
    ("NADRA verification: OTP code share karein jo aap ko mila hai. Yeh government requirement hai.", "Mixed", "OTP", "Scam"),
    ("Your account has been temporarily locked. Enter the OTP sent to your phone to unlock immediately.", "English", "OTP", "Scam"),
    ("آپ کی آن لائن خریداری مکمل کرنے کے لیے OTP درکار ہے۔ براہ کرم کوڈ فراہم کریں۔", "Urdu", "OTP", "Scam"),

    # ──────────────── ADDITIONAL SIM BLOCK SCAMS ────────────────
    ("NOTICE: Aap ka Jazz SIM agle 6 ghante mein permanently band ho jaye ga. Bachane ke liye Rs. 500 bhejein.", "Roman Urdu", "SIM Block", "Scam"),
    ("Telenor Pakistan: Your SIM has been reported stolen. Reactivate by sending CNIC details and Rs. 1,000.", "English", "SIM Block", "Scam"),
    ("آپ کے موبائل نمبر پر غیر قانونی کالز کا پتہ چلا ہے۔ PTA نے بلاک کرنے کا حکم دیا ہے۔", "Urdu", "SIM Block", "Scam"),
    ("Aap ka Zong connection terminate hone wala hai. Rs. 2,000 pay karke reactivate karein.", "Roman Urdu", "SIM Block", "Scam"),
    ("Government order: All SIMs without updated CNIC will be blocked by Friday. Upload CNIC at http://sim-verify.pk", "English", "SIM Block", "Scam"),
    ("PTA se final warning: Aap ke 3 SIMs block hone wali hain. Sirf Rs. 200 verification fee se bachayein.", "Mixed", "SIM Block", "Scam"),
    ("Ufone user alert: Your number is being transferred to another person. Block this by calling and providing CNIC.", "English", "SIM Block", "Scam"),
    ("Aap ka number blacklist mein daal diya gaya hai. Remove karne ke liye is number par Rs. 1,500 easyload bhejein.", "Roman Urdu", "SIM Block", "Scam"),
    ("SIM verification scam: Apna fingerprint scan karwao nearest E-Sahulat center par. Rs. 300 charges hain.", "Mixed", "SIM Block", "Scam"),
    ("آپ کا سم کارڈ غیر فعال ہو چکا ہے۔ بحالی کے لیے Rs. 800 اس نمبر پر بھیجیں۔", "Urdu", "SIM Block", "Scam"),

    # ──────────────── ADDITIONAL INVESTMENT SCAMS ────────────────
    ("MLM opportunity: Join Rs. 10,000 mein aur 3 logon ko refer karke Rs. 50,000 kamayein. Pyramid... I mean network marketing!", "Mixed", "Investment", "Scam"),
    ("Real estate plot booking: Rs. 25,000 token money dein aur DHA mein 5 marla plot book karein. Limited offer!", "Mixed", "Investment", "Scam"),
    ("آپ صرف Rs. 2,000 روزانہ بچائیں اور 6 ماہ میں کروڑ پتی بنیں۔ حیرت انگیز اسکیم!", "Urdu", "Investment", "Scam"),
    ("Online trading academy: Learn forex in 7 days! Course fee Rs. 20,000. Guaranteed $500 daily profit after training.", "English", "Investment", "Scam"),
    ("Solar panel business opportunity! Rs. 50,000 invest karein aur monthly Rs. 25,000 profit earn karein lifetime.", "Mixed", "Investment", "Scam"),
    ("Dubai gold trading: Rs. 1,00,000 invest karein, monthly 30% return guaranteed. Dubai office se receipt milegi.", "Mixed", "Investment", "Scam"),
    ("Artificial intelligence trading bot! Rs. 15,000 se start karein aur daily Rs. 2,000 automatic profit kamayein.", "Mixed", "Investment", "Scam"),
    ("کسان بھائیوں کے لیے زرعی سرمایہ کاری: Rs. 30,000 لگائیں اور فصل سے دگنا منافع کمائیں۔", "Urdu", "Investment", "Scam"),
    ("Uber fleet owner scheme: Rs. 2,00,000 invest karo, hum car khareed ke chalayein ge. Monthly Rs. 40,000 profit.", "Roman Urdu", "Investment", "Scam"),
   ("App development mein invest karein! Rs. 25,000 dalein aur app ki revenue se monthly 50% share lein.", "Mixed", "Investment", "Scam"),

    # ──────────────── IMPERSONATION SCAMS (EXPANDED - subtle patterns) ────────────────
    ("Assalam o Alaikum, yeh Ammi ka naya number hai. Phone gir gaya tha. Beta Rs. 15,000 bhej do, dawai leni hai.", "Roman Urdu", "Impersonation", "Scam"),
    ("Bhai, main airport par hoon. Flight miss ho gayi. Rs. 25,000 bhejo nayi ticket ke liye. Ghar aa ke bataonga.", "Roman Urdu", "Impersonation", "Scam"),
    ("Salam bhai. Main hoon na, aap ka purana classmate Ahmed. Ek emergency hai, Rs. 8,000 chahiye. Kal mil ke baat karte hain.", "Roman Urdu", "Impersonation", "Scam"),
    ("Beta main Papa hoon. Meeting mein hoon, phone kharab hai. Ek kaam karo - Rs. 35,000 is account mein transfer karo, boss ka account hai.", "Roman Urdu", "Impersonation", "Scam"),
    ("Hey, this is your colleague Sara from the Lahore office. I'm stuck at a client meeting and need Rs. 5,000 for cab fare. Will reimburse tomorrow.", "English", "Impersonation", "Scam"),
    ("Aap ki phupho bol rahi hoon. Bete ka accident ho gaya hai. Rs. 50,000 hospital mein deposit karne hain. Jaldi bhejein.", "Roman Urdu", "Impersonation", "Scam"),
    ("Dear, main office trip pe hoon. Wallet kho gaya. Is colleague ke account mein Rs. 20,000 bhej do. Baad mein explain karti hoon.", "Roman Urdu", "Impersonation", "Scam"),
    ("Yeh aap ka chacha hoon. Dubai se call kar raha hoon. Passport kho gaya hai, Rs. 40,000 chahiye emergency travel document ke liye.", "Roman Urdu", "Impersonation", "Scam"),
    ("Bhai jan, main hoon Tariq. Yaar ek masla ho gaya - car accident. Rs. 30,000 chahiye towing aur repair ke liye. Aaj raat tak wapis.", "Roman Urdu", "Impersonation", "Scam"),
    ("This is Mr. Khan from your company's HR department. We need to process your salary early. Please confirm your bank account and password for verification.", "English", "Impersonation", "Scam"),
    ("Aslam o Alaikum bhai. Main aap ka padosi Imran hoon. Ghar mein emergency hai, biwi ko hospital le jana hai. Rs. 10,000 udhaar chahiye.", "Roman Urdu", "Impersonation", "Scam"),
    ("Mom here - phone battery dying. At the hospital with Dad. He needs immediate tests costing Rs. 15,000. Please send to this number ASAP.", "English", "Impersonation", "Scam"),
    ("Bhai, yeh mera backup number hai. Asli phone paani mein gir gaya. Mujhe Rs. 12,000 send karo, naya phone lena hai zaroori kaam ke liye.", "Roman Urdu", "Impersonation", "Scam"),
    ("Salam Alaikum. Main aap ka dost Rashid hoon, Canada se. Mera Pakistani account freeze ho gaya. Rs. 25,000 bhejo, wapis aake chukaonga.", "Roman Urdu", "Impersonation", "Scam"),
    ("Boss here. I'm in an important meeting with investors. Transfer Rs. 75,000 to this vendor account immediately. I'll explain after the meeting.", "English", "Impersonation", "Scam"),
    ("Beti, main tumhari khala hoon. Naya number save kar lo. Ek zaroori kaam hai - Rs. 8,000 bhej do, bachon ki school fee deni hai.", "Roman Urdu", "Impersonation", "Scam"),
    ("Yaar main Farhan hoon. Bhai ek help chahiye - Rs. 5,000 JazzCash pe bhej do. Phone band hone wala hai, baad mein call karta hoon.", "Roman Urdu", "Impersonation", "Scam"),
    ("This is your landlord. I need Rs. 25,000 urgently for property tax payment. Please transfer to my new account number. Receipt will be provided.", "English", "Impersonation", "Scam"),
    ("Aap ki bhabi hoon. Bhai sahab ka phone switch off hai. Unki tabiyat kharab hai, Rs. 20,000 hospital ke liye bhej dein.", "Roman Urdu", "Impersonation", "Scam"),
    ("Salam, main aap ka bhanja hoon - Saad. Mama ka naya number hai yeh. Mujhe university fee ke liye Rs. 30,000 chahiye. Papa ne kaha aap se loon.", "Roman Urdu", "Impersonation", "Scam"),
    ("Hi, this is Ayesha from your gym. The trainer needs Rs. 3,000 advance for your personal training package. Send to this number to confirm your slot.", "English", "Impersonation", "Scam"),
    ("Bhai, main ghar se baat kar raha hoon. Ammi ki tabiyat theek nahi, doctor ne test karwaye hain. Rs. 7,000 lab fee ke liye bhejo.", "Roman Urdu", "Impersonation", "Scam"),
    ("آپ کے بھائی کا نیا نمبر۔ فون ٹوٹ گیا ہے۔ Rs. 18,000 بھیجو، اگلے ہفتے واپس کروں گا۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کی بہن کا پیغام: بھائی مجھے بچوں کی فیس کے لیے Rs. 12,000 چاہیے۔ یہ نیا نمبر ہے۔", "Urdu", "Impersonation", "Scam"),
    ("السلام علیکم! میں آپ کا ماموں ہوں۔ بیٹے کی شادی ہے، Rs. 50,000 ادھار چاہیے۔ بعد میں واپس کروں گا۔", "Urdu", "Impersonation", "Scam"),
    ("Office colleague here: Your team lead asked me to collect Rs. 10,000 from each member for the surprise party. Send to my JazzCash.", "English", "Impersonation", "Scam"),
    ("Beta main Ammi hoon, naye phone se message kar rahi hoon. Papa ka phone bhi band hai. Rs. 6,000 bhejo, grocery leni hai.", "Roman Urdu", "Impersonation", "Scam"),
    ("Bhaijaan, main Dubai airport pe hoon. Visa expire ho raha hai, Rs. 35,000 fine pay karna hai warna deport ho jaonga. Foran bhejein.", "Roman Urdu", "Impersonation", "Scam"),
    ("This is your child's school principal. We need immediate payment of Rs. 8,000 for the annual function. Transfer to this account to confirm participation.", "English", "Impersonation", "Scam"),
    ("Assalam o Alaikum, yeh masjid committee se hain. Ramadan fundraiser ke liye Rs. 5,000 is account par bhejein. Receipt di jaye gi.", "Roman Urdu", "Impersonation", "Scam"),
    ("Hi this is your neighbor Kamran. Bhai meri car kharab ho gayi hai aur tow truck wale ko Rs. 4,000 chahiye. Udhaar de do kal wapis.", "Roman Urdu", "Impersonation", "Scam"),
    ("Bhai, main hoon - aap ka cousin jo Karachi mein rehta hai. Number change ho gaya. Ek emergency hai, Rs. 15,000 bhej do please.", "Roman Urdu", "Impersonation", "Scam"),
    ("آپ کے آفس کا عملہ: آپ کی تنخواہ روک دی گئی ہے۔ فوری طور پر تصدیقی فیس Rs. 3,000 جمع کروائیں۔", "Urdu", "Impersonation", "Scam"),
    ("Sister here - mera phone kho gaya. Yeh mera naya number hai. Mujhe Rs. 10,000 send karo, bachon ki books leni hain.", "Roman Urdu", "Impersonation", "Scam"),
    ("Papa here - stuck in Islamabad. Car broke down on motorway. Need Rs. 20,000 for repairs. Send to mechanic's account. Will call when fixed.", "English", "Impersonation", "Scam"),

    # ──────────────── JOB SCAMS (EXPANDED) ────────────────
    ("LinkedIn profile selected! Aap ko Fortune 500 company se offer hai. Sirf Rs. 8,000 documentation fee pay karein.", "Mixed", "Job", "Scam"),
    ("Freelance writing job: Earn $500/week writing articles. Security deposit $50 required. No experience needed!", "English", "Job", "Scam"),
    ("Aap ki qualification dekh kar hum bohat impress hain. Rs. 25,000 monthly salary guarantee. Training fee Rs. 3,500 advance mein.", "Roman Urdu", "Job", "Scam"),
    ("Govt of Pakistan job portal: Aap ka naam shortlist hua hai. Interview fee Rs. 5,000 pay karein is link par.", "Mixed", "Job", "Scam"),
    ("Uber driver partnership: Rs. 2,00,000 invest karein, car hum provide karein ge. Monthly Rs. 60,000 income guaranteed.", "Mixed", "Job", "Scam"),
    ("Night shift call center job - salary Rs. 45,000. Security deposit Rs. 4,000 required. Start immediately after payment.", "English", "Job", "Scam"),
    ("آپ کو سعودی عرب میں پلمبر کی نوکری مل سکتی ہے۔ ایجنسی فیس Rs. 20,000۔ ویزا 10 دن میں۔", "Urdu", "Job", "Scam"),
    ("WhatsApp business group join karein aur daily Rs. 3,000 kamayein by liking YouTube videos. Registration Rs. 2,000.", "Mixed", "Job", "Scam"),
    ("Pharmacy assistant job in Canada - salary CAD $4000/month. Pay Rs. 45,000 visa processing fee. Guaranteed approval.", "English", "Job", "Scam"),
    ("Bhai, ek construction company mein 100 mazdoor chahiye. Daily wages Rs. 2,500. Registration Rs. 1,500 per person.", "Roman Urdu", "Job", "Scam"),

    # ──────────────── BANK FRAUD (EXPANDED) ────────────────
    ("HBL Security Alert: Your account was accessed from an unrecognized device in Islamabad. Confirm your identity: http://hbl-confirm.net", "English", "Bank", "Scam"),
    ("Aap ka Meezan Bank account mein Rs. 3,50,000 ki suspicious transaction detect hui hai. Cancel karne ke liye yeh link kholen.", "Roman Urdu", "Bank", "Scam"),
    ("آپ کا یو بی ایل اکاؤنٹ 48 گھنٹے میں بند ہو جائے گا۔ فوری تصدیق کے لیے یہ لنک استعمال کریں۔", "Urdu", "Bank", "Scam"),
    ("Your JazzCash has been upgraded to premium. Annual charges Rs. 2,500 will apply. Cancel by visiting http://jazzcash-premium.com", "English", "Bank", "Scam"),
    ("SBI Pakistan branch: Your locker needs verification. Bring CNIC and share locker PIN for annual audit compliance.", "English", "Bank", "Scam"),
    ("Aap ki credit card se Dubai mein Rs. 85,000 ki shopping hui hai. Agar aap ne nahi ki to immediately yeh form fill karein.", "Roman Urdu", "Bank", "Scam"),
    ("Nayapay Alert: Your account has been flagged for suspicious transfers. Verify at http://nayapay-verify.com within 24 hours.", "English", "Bank", "Scam"),
    ("SadaPay: Your account is under review for potential money laundering. Share CNIC front/back photos and transaction history.", "English", "Bank", "Scam"),
    ("آپ کے اکاؤنٹ سے بیرون ملک ٹرانزیکشن کی کوشش ناکام۔ فوری طور پر کارڈ نمبر اور CVV اپ ڈیٹ کریں۔", "Urdu", "Bank", "Scam"),
    ("Bank Al Habib: Aap ki cheque book dispatch ready hai. Rs. 500 courier charges is link par pay karein.", "Mixed", "Bank", "Scam"),

    # ──────────────── INVESTMENT SCAMS (EXPANDED) ────────────────
    ("PSX certified broker: Rs. 50,000 invest karein stocks mein aur monthly 25% return payein. SECP registration number attached.", "Mixed", "Investment", "Scam"),
    ("Bitcoin Pakistan: Rs. 10,000 se crypto trading shuru karein. Daily Rs. 2,000 profit guaranteed. Withdraw anytime!", "Mixed", "Investment", "Scam"),
    ("آپ صرف Rs. 3,000 روزانہ بچائیں اور 3 ماہ میں Rs. 5,00,000 کمائیں۔ حیرت انگیز سرمایہ کاری!", "Urdu", "Investment", "Scam"),
    ("Daraz seller program: Rs. 20,000 invest karein inventory mein aur monthly Rs. 15,000 profit kamayein. No risk!", "Mixed", "Investment", "Scam"),
    ("YouTube automation course: Rs. 30,000 pay karein aur hum aap ka channel auto-generate karein ge. Monthly $1000 income guaranteed.", "Mixed", "Investment", "Scam"),
    ("TikTok earning app: Rs. 5,000 deposit karein aur videos dekh kar daily Rs. 1,500 kamayein. 10,000 users already earning!", "Mixed", "Investment", "Scam"),
    ("Islamic halal investment: Rs. 25,000 se shuru karein aur monthly 15% munafa payein. Shariah compliant guaranteed!", "Roman Urdu", "Investment", "Scam"),
    ("Property dealer: Rs. 1,00,000 token money dein aur Bahria Town mein 10 marla plot book karein. Files ready, 100% genuine.", "Mixed", "Investment", "Scam"),
    ("Dropshipping business: Rs. 15,000 invest karein aur Shopify store hum banayein ge. Monthly Rs. 50,000 revenue guaranteed.", "Mixed", "Investment", "Scam"),
    ("AI trading bot Pakistan: Rs. 8,000 se start karein. AI khud trade karega aur daily 5% profit dega. No loss guarantee!", "Mixed", "Investment", "Scam"),

    # ──────────────── PRIZE SCAMS (EXPANDED) ────────────────
    ("Daraz 11.11 Lucky Winner! Aap ne Rs. 50,000 ka voucher jeeta hai. Claim fee Rs. 1,500 is link par pay karein.", "Mixed", "Prize", "Scam"),
    ("Telenor 25th Anniversary: Aap ko free iPhone 14 mila hai! Delivery charges Rs. 3,500 advance pay karein.", "Mixed", "Prize", "Scam"),
    ("You've WON a free Umrah ticket from Pakistan International Airlines! Pay Rs. 25,000 visa processing fee to claim.", "English", "Prize", "Scam"),
    ("Coca-Cola cap prize: Code match ho gaya! Rs. 5,00,000 ka cash prize. Claim karne ke liye Rs. 10,000 bhejein.", "Mixed", "Prize", "Scam"),
    ("JazzCash 10 saal poore: Lucky draw mein aap ka naam aaya hai - Rs. 10,00,000 prize. Processing fee Rs. 20,000.", "Mixed", "Prize", "Scam"),
    ("Samsung Pakistan giveaway: Aap ne Galaxy S24 Ultra jeeta hai! Shipping fee Rs. 5,000 is account par bhejein.", "Mixed", "Prize", "Scam"),
    ("PTV license fee refund: Aap ko Rs. 3,000 wapis milne hain. Processing ke liye Rs. 500 advance bhejein.", "Mixed", "Prize", "Scam"),
    ("آپ نے آن لائن مقابلے میں نئی گاڑی جیتی ہے! Rs. 50,000 ڈلیوری اور رجسٹریشن فیس ادا کریں۔", "Urdu", "Prize", "Scam"),
    ("FoodPanda anniversary: Aap ko Rs. 25,000 ka free food voucher mila hai! Activation fee Rs. 800 pay karein.", "Mixed", "Prize", "Scam"),
    ("Easypaisa Mega Reward: Your number has been selected for Rs. 20,00,000 cash prize. Transfer Rs. 15,000 to claim.", "English", "Prize", "Scam"),

    # ──────────────── OTP SCAMS (EXPANDED) ────────────────
    ("Customer care se baat ho rahi hai: Sir aap ka Rs. 50,000 ka refund process karne ke liye OTP chahiye jo abhi aap ko SMS pe aaya.", "Roman Urdu", "OTP", "Scam"),
    ("JazzCash helpline: Aap ka account suspend hone wala hai. Reactivate karne ke liye jo code aaya hai woh batayein.", "Roman Urdu", "OTP", "Scam"),
    ("Bank security team: We detected unauthorized access. Share the OTP sent to your phone to lock your account immediately.", "English", "OTP", "Scam"),
    ("Aap ki online shopping ki delivery confirm karne ke liye OTP chahiye. Jo SMS pe code aaya hai woh rider ko batayein.", "Roman Urdu", "OTP", "Scam"),
    ("EasyPaisa reward: Rs. 10,000 cashback paane ke liye apna OTP code share karein. Offer sirf 30 minute ke liye!", "Roman Urdu", "OTP", "Scam"),
    ("آپ کے اکاؤنٹ کی سیکیورٹی اپ ڈیٹ کے لیے OTP کوڈ درکار ہے۔ فوری طور پر کوڈ شیئر کریں۔", "Urdu", "OTP", "Scam"),
    ("Technical support calling: Sir aap ka phone hack ho raha hai. Jo verification code aaya hai woh humein dein, hum secure karein ge.", "Roman Urdu", "OTP", "Scam"),
    ("NADRA biometric update: OTP code share karein jo aap ko SMS pe mila. Yeh mandatory government requirement hai.", "Mixed", "OTP", "Scam"),

    # ──────────────── SIM BLOCK SCAMS (EXPANDED) ────────────────
    ("Jazz: Aap ka number kal sham tak band ho jaye ga. Bachane ke liye Rs. 1,000 is special number par recharge karein.", "Roman Urdu", "SIM Block", "Scam"),
    ("PTA MANDATORY: All SIMs must be re-verified with CNIC by Friday. Upload at http://pta-sim-verify.pk or pay Rs. 500 fine.", "English", "SIM Block", "Scam"),
    ("آپ کا ٹیلی نار سم کارڈ غیر فعال ہونے والا ہے۔ Rs. 800 فیس ادا کر کے بحال کریں۔", "Urdu", "SIM Block", "Scam"),
    ("Zong: Aap ke naam par 8 SIMs registered hain jo illegal hain. Rs. 2,000 verification fee bhejein warna police complaint hogi.", "Roman Urdu", "SIM Block", "Scam"),
    ("Ufone security notice: Your SIM has been cloned. Pay Rs. 1,500 for SIM replacement or your number will be given to someone else.", "English", "SIM Block", "Scam"),
    ("Telenor Pakistan: Aap ka postpaid connection disconnect hone wala hai. Rs. 3,000 outstanding bill + Rs. 500 reconnection fee abhi pay karein.", "Roman Urdu", "SIM Block", "Scam"),
    ("Govt notification: SIM cards without NADRA verification will be permanently blocked. Rs. 200 verification fee required.", "English", "SIM Block", "Scam"),
    ("آپ کا سم کارڈ دوسرے کے نام پر ٹرانسفر ہو رہا ہے۔ فوری طور پر Rs. 1,200 بھیج کر روکیں۔", "Urdu", "SIM Block", "Scam"),

    # ──────────────── LOTTERY SCAMS (EXPANDED) ────────────────
    ("Jazz Cash mega lucky draw: Aap ka number Rs. 50,00,000 ka winner hai! Sirf Rs. 12,000 tax bhej ke claim karein.", "Mixed", "Lottery", "Scam"),
    ("WhatsApp lottery winner! Aap ne forward chain complete kiya - Rs. 2,00,000 prize. Claim fee Rs. 5,000.", "Mixed", "Lottery", "Scam"),
    ("Turkey tourism board lucky draw: You won a free trip to Istanbul! Pay Rs. 35,000 for visa and flight booking charges.", "English", "Lottery", "Scam"),
    ("آپ کا موبائل نمبر رمضان قرعہ اندازی میں Rs. 15 لاکھ کا انعام جیتا ہے۔ Rs. 20,000 ٹیکس ادا کریں۔", "Urdu", "Lottery", "Scam"),
    ("Pepsi bottle cap code winner! Rs. 5,00,000 cash prize. Processing fee Rs. 8,000 bhejein warna prize cancel ho jaye ga.", "Mixed", "Lottery", "Scam"),
    ("Imran Khan Foundation lucky draw: Aap ko Rs. 10,00,000 milein ge. Rs. 15,000 donation fee pay karein pehle.", "Mixed", "Lottery", "Scam"),
    ("Eid Mubarak special draw! First prize: Rs. 50,00,000. Ticket price Rs. 1,000 only. Buy now before draw closes tonight!", "English", "Lottery", "Scam"),
    ("National Bond Pakistan: Aap ka bond number winner hai! Rs. 25,00,000 prize. Rs. 30,000 processing fee advance mein.", "Mixed", "Lottery", "Scam"),

    # ──────────────── OTP SCAMS (URDU - targeting FN pattern) ────────────────
    ("آپ کی آن لائن خریداری مکمل کرنے کے لیے OTP درکار ہے۔ براہ کرم کوڈ فراہم کریں۔", "Urdu", "OTP", "Scam"),
    ("آپ کا اکاؤنٹ محفوظ کرنے کے لیے OTP کوڈ ہمیں بتائیں۔ ہم آپ کا اکاؤنٹ محفوظ کریں گے۔", "Urdu", "OTP", "Scam"),
    ("براہ کرم اپنا تصدیقی کوڈ فراہم کریں تاکہ آپ کی رقم منتقلی مکمل ہو سکے۔", "Urdu", "OTP", "Scam"),
    ("آپ کو بھیجا گیا کوڈ ہمیں بتائیں، بینک کی طرف سے تصدیق ضروری ہے۔", "Urdu", "OTP", "Scam"),
    ("آپ کے اکاؤنٹ کی تصدیق کے لیے OTP کوڈ درکار ہے۔ SMS پر موصول ہونے والا کوڈ بھیجیں۔", "Urdu", "OTP", "Scam"),
    ("آپ کا اکاؤنٹ ہیک ہو رہا ہے۔ فوری طور پر OTP کوڈ شیئر کریں تاکہ ہم محفوظ کر سکیں۔", "Urdu", "OTP", "Scam"),
    ("آپ کی رقم کی منتقلی روک دی گئی ہے۔ OTP کوڈ فراہم کریں تاکہ منتقلی مکمل ہو۔", "Urdu", "OTP", "Scam"),
    ("آپ کا جیز کیش اکاؤنٹ بند ہونے والا ہے۔ OTP کوڈ شیئر کریں ورنہ اکاؤنٹ بند ہو جائے گا۔", "Urdu", "OTP", "Scam"),

    # ──────────────── URDU JOB SCAMS ────────────────
    ("آپ کو متحدہ عرب امارات میں اعلیٰ تنخواہ والی نوکری کی پیشکش ہے۔ Rs. 15,000 رجسٹریشن فیس ادا کریں۔", "Urdu", "Job", "Scam"),
    ("گھر بیٹھے Rs. 60,000 ماہانہ کمائیں۔ صرف Rs. 3,000 رجسٹریشن فیس درکار ہے۔", "Urdu", "Job", "Scam"),
    ("آپ کا سی وی منتخب ہو گیا ہے۔ قطر میں ہوٹل کی نوکری۔ ایجنسی فیس Rs. 25,000۔", "Urdu", "Job", "Scam"),
    ("فوری بھرتی! سرکاری محکمے میں 50 آسامیاں خالی ہیں۔ Rs. 10,000 فیس جمع کروائیں۔", "Urdu", "Job", "Scam"),
    ("آپ کو کینیڈا میں امیگریشن کی پیشکش ہے۔ Rs. 50,000 پروسیسنگ فیس ادا کریں۔", "Urdu", "Job", "Scam"),
    ("آن لائن ٹائپنگ جاب: روزانہ 10 صفحات ٹائپ کریں اور Rs. 3,000 فی صفحہ کمائیں۔ سکیورٹی ڈپازٹ Rs. 5,000۔", "Urdu", "Job", "Scam"),
    ("آپ کو ملائیشیا میں فیکٹری ورکر کی نوکری مل سکتی ہے۔ Rs. 20,000 فیس۔ ویزا 15 دن میں۔", "Urdu", "Job", "Scam"),
    ("پارٹ ٹائم ڈیٹا انٹری جاب: روزانہ 2 گھنٹے کام کریں اور Rs. 50,000 ماہانہ کمائیں۔ رجسٹریشن Rs. 4,000۔", "Urdu", "Job", "Scam"),
    ("آپ کی اہلیت دیکھ کر ہم بہت متاثر ہیں۔ Rs. 40,000 ماہانہ تنخواہ کی ضمانت۔ ٹریننگ فیس Rs. 5,000۔", "Urdu", "Job", "Scam"),
    ("سعودی عرب میں الیکٹریشن کی 30 آسامیاں۔ Rs. 18,000 رجسٹریشن فیس۔ محدود نشستیں۔", "Urdu", "Job", "Scam"),
    ("آپ کو عمان میں ڈرائیور کی نوکری مل سکتی ہے۔ ویزا فیس Rs. 22,000۔ آج ہی درخواست دیں۔", "Urdu", "Job", "Scam"),
    ("فری لانس رائٹنگ جاب: مضامین لکھ کر Rs. 5,000 فی مضمون کمائیں۔ سکیورٹی ڈپازٹ Rs. 2,000۔", "Urdu", "Job", "Scam"),

    # ──────────────── URDU LOTTERY SCAMS ────────────────
    ("مبارک ہو! آپ نے Rs. 75,00,000 کی قرعہ اندازی جیت لی ہے۔ کلیم کرنے کے لیے Rs. 15,000 ٹیکس ادا کریں۔", "Urdu", "Lottery", "Scam"),
    ("آپ کا موبائل نمبر بین الاقوامی لاٹری میں منتخب ہوا ہے۔ $200,000 انعام۔ پروسیسنگ فیس $100۔", "Urdu", "Lottery", "Scam"),
    ("آپ نے Rs. 25,00,000 کا نقد انعام جیتا ہے! کلیم کرنے کے لیے Rs. 20,000 فیس بھیجیں۔", "Urdu", "Lottery", "Scam"),
    ("پیپسی کیپ کوڈ مقابلہ: آپ نے Rs. 10,00,000 جیتے ہیں! Rs. 12,000 پروسیسنگ فیس ادا کریں۔", "Urdu", "Lottery", "Scam"),
    ("آپ کا نمبر رمضان مبارک قرعہ اندازی میں Rs. 50,00,000 کا فاتح ہے۔ Rs. 25,000 ٹیکس ادا کریں۔", "Urdu", "Lottery", "Scam"),
    ("عید اسپیشل ڈرا: پہلا انعام Rs. 1 کروڑ۔ ٹکٹ صرف Rs. 500۔ آج ہی خریدیں!", "Urdu", "Lottery", "Scam"),
    ("آپ نے ہاؤسنگ سوسائٹی کے ڈرا میں 10 مرلہ پلاٹ جیتا ہے۔ Rs. 50,000 ٹرانسفر فیس ادا کریں۔", "Urdu", "Lottery", "Scam"),
    ("ٹیلی نار لکی ڈرا: آپ کا نمبر Rs. 30,00,000 کا فاتح ہے۔ کلیم فیس Rs. 18,000۔", "Urdu", "Lottery", "Scam"),

    # ──────────────── URDU BANK FRAUD ────────────────
    ("آپ کا حبیب بینک اکاؤنٹ غیر مجاز لاگ ان سے متاثر ہوا ہے۔ فوری تصدیق کے لیے یہ لنک کھولیں۔", "Urdu", "Bank", "Scam"),
    ("آپ کے میزان بینک اکاؤنٹ سے Rs. 2,50,000 کی مشکوک ٹرانزیکشن ہوئی ہے۔ منسوخ کرنے کے لیے لنک پر کلک کریں۔", "Urdu", "Bank", "Scam"),
    ("آپ کا ڈیبٹ کارڈ بلاک کر دیا گیا ہے۔ ان بلاک کرنے کے لیے کال کریں اور اپنی تمام تفصیلات بتائیں۔", "Urdu", "Bank", "Scam"),
    ("اسٹیٹ بینک: آپ کا اکاؤنٹ فلैग ہو گیا ہے۔ CNIC اور بینک پاس ورڈ جمع کروائیں۔", "Urdu", "Bank", "Scam"),
    ("آپ کے اکاؤنٹ سے بیرون ملک Rs. 1,20,000 کی ٹرانزیکشن کی کوشش ناکام۔ فوری تصدیق کریں۔", "Urdu", "Bank", "Scam"),
    ("ایف بی آر ٹیکس ریفنڈ: آپ کو Rs. 95,000 کا ریفنڈ ملا ہے۔ بینک تفصیلات اس لنک پر درج کریں۔", "Urdu", "Bank", "Scam"),
    ("آپ کا نیّا پے اکاؤنٹ مشکوک سرگرمی کی وجہ سے فلैग ہو گیا ہے۔ 24 گھنٹے میں تصدیق کریں۔", "Urdu", "Bank", "Scam"),
    ("آپ کے کریڈٹ کارڈ سے دبئی میں Rs. 65,000 کی خریداری ہوئی ہے۔ اگر آپ نے نہیں کی تو یہ فارم پُر کریں۔", "Urdu", "Bank", "Scam"),
    ("بینک الفلاح: آپ کی چیک بک ڈسپیچ تیار ہے۔ Rs. 500 کوریئر چارجز اس لنک پر ادا کریں۔", "Urdu", "Bank", "Scam"),
    ("آپ کے اکاؤنٹ کی کے وائی سی تصدیق نہیں ہوئی۔ 48 گھنٹے میں اکاؤنٹ منجمد ہو جائے گا۔ فوری عمل کریں۔", "Urdu", "Bank", "Scam"),

    # ──────────────── URDU INVESTMENT SCAMS ────────────────
    ("آپ صرف Rs. 5,000 سرمایہ کاری کریں اور ماہانہ 200% منافع کمائیں۔ یہ موقع دوبارہ نہیں آئے گا!", "Urdu", "Investment", "Scam"),
    ("بٹ کوائن پاکستان: Rs. 10,000 سے کرپٹو ٹریڈنگ شروع کریں۔ روزانہ Rs. 3,000 منافع کی ضمانت۔", "Urdu", "Investment", "Scam"),
    ("آئی ٹی سیکٹر میں سرمایہ کاری: Rs. 50,000 لگائیں اور سالانہ 300% ریٹرن حاصل کریں۔", "Urdu", "Investment", "Scam"),
    ("پراپرٹی ڈیلر: Rs. 1,00,000 ٹوکن منی دیں اور بحریہ ٹاؤن میں 10 مرلہ پلاٹ بک کریں۔ فائلیں تیار۔", "Urdu", "Investment", "Scam"),
    ("ڈراپ شپنگ بزنس: Rs. 15,000 سرمایہ کاری کریں۔ ماہانہ Rs. 50,000 ریونیو کی ضمانت۔", "Urdu", "Investment", "Scam"),
    ("سولر پینل بزنس: Rs. 50,000 لگائیں اور ماہانہ Rs. 25,000 منافع کمائیں۔ تاحیات ضمانت!", "Urdu", "Investment", "Scam"),
    ("کسان بھائیوں کے لیے زرعی سرمایہ کاری: Rs. 30,000 لگائیں اور فصل سے دگنا منافع کمائیں۔", "Urdu", "Investment", "Scam"),
    ("اے آئی ٹریڈنگ بوٹ: Rs. 8,000 سے شروع کریں۔ بوٹ خود ٹریڈ کرے گا اور روزانہ 5% منافع دے گا۔", "Urdu", "Investment", "Scam"),
    ("یوٹیوب آٹومیشن کورس: Rs. 30,000 ادا کریں۔ ماہانہ $1000 آمدنی کی ضمانت۔", "Urdu", "Investment", "Scam"),
    ("حلال سرمایہ کاری: Rs. 25,000 سے شروع کریں اور ماہانہ 15% منافع پائیں۔ شریعت کے مطابق!", "Urdu", "Investment", "Scam"),

    # ──────────────── URDU PRIZE SCAMS ────────────────
    ("آپ نے آئی فون 15 پرو جیت لیا ہے! ڈلیوری چارجز Rs. 2,500 ادا کریں۔", "Urdu", "Prize", "Scam"),
    ("آپ نے قرعہ اندازی میں نئی ہونڈا سِوک جیتی ہے! Rs. 25,000 ڈلیوری چارجز ادا کریں۔", "Urdu", "Prize", "Scam"),
    ("داراز 11.11 لکی ونر! آپ نے Rs. 50,000 کا واؤچر جیتا ہے۔ کلیم فیس Rs. 1,500۔", "Urdu", "Prize", "Scam"),
    ("آپ نے عمرہ پیکج جیت لیا ہے! صرف Rs. 50,000 ویزا اور ٹکٹ چارجز ادا کریں۔", "Urdu", "Prize", "Scam"),
    ("سیمسنگ پاکستان گیو اوے: آپ نے گلیکسی S24 الٹرا جیتا ہے! شپنگ فیس Rs. 5,000۔", "Urdu", "Prize", "Scam"),
    ("فلائی اسکول وفاداری انعام: آپ کو Rs. 10,000 کا واؤچر ملا ہے! ایکٹیویشن فیس Rs. 300۔", "Urdu", "Prize", "Scam"),
    ("آپ نے آن لائن مقابلے میں Rs. 1,00,000 کا شاپنگ واؤچر جیتا ہے۔ Rs. 2,000 فیس ادا کریں۔", "Urdu", "Prize", "Scam"),
    ("آپ نے شاپنگ سپری مقابلے میں نئی گاڑی جیتی ہے! Rs. 50,000 ڈلیوری اور رجسٹریشن فیس ادا کریں۔", "Urdu", "Prize", "Scam"),

    # ──────────────── URDU SIM BLOCK SCAMS ────────────────
    ("آپ کا جاز سم کارڈ 24 گھنٹے میں بلاک ہو جائے گا۔ بچانے کے لیے Rs. 1,000 بھیجیں۔", "Urdu", "SIM Block", "Scam"),
    ("آپ کے نام پر 5 سم کارڈز درج ہیں جو غیر قانونی ہیں۔ Rs. 2,000 تصدیقی فیس بھیجیں ورنہ پولیس شکایت ہوگی۔", "Urdu", "SIM Block", "Scam"),
    ("پی ٹی اے نوٹس: آپ کا موبائل نمبر نادرا سے تصدیق نہیں ہوا۔ CNIC تفصیلات شیئر کریں ورنہ سم بلاک۔", "Urdu", "SIM Block", "Scam"),
    ("ٹیلی نار: آپ کا سم کلون ہو گیا ہے۔ Rs. 1,500 ادا کریں ورنہ نمبر کسی اور کو دیا جائے گا۔", "Urdu", "SIM Block", "Scam"),
    ("حکومتی حکم: بغیر تصدیق کے تمام سم کارڈز جمعے تک بلاک ہو جائیں گے۔ Rs. 200 تصدیقی فیس درکار۔", "Urdu", "SIM Block", "Scam"),
    ("یوفون: آپ کا سم کارڈ دوسرے کے نام پر ٹرانسفر ہو رہا ہے۔ فوری طور پر Rs. 1,200 بھیج کر روکیں۔", "Urdu", "SIM Block", "Scam"),
    ("آپ کا زونگ سم کارڈ غیر فعال ہونے والا ہے۔ Rs. 800 فیس ادا کر کے بحال کریں۔", "Urdu", "SIM Block", "Scam"),
    ("آپ کا سم کارڈ غیر قانونی درآمد سے متعلق ہے۔ پی ٹی اے دفتر میں رپورٹ کریں یا CNIC بھیج کر کلیئر کریں۔", "Urdu", "SIM Block", "Scam"),

    # ──────────────── URDU IMPERSONATION SCAMS ────────────────
    ("السلام علیکم! میں آپ کا بھائی ہوں۔ نیا نمبر ہے۔ مجھے فوری Rs. 15,000 کی ضرورت ہے۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کے دوست کا نیا نمبر۔ مجھے ابھی Rs. 8,000 بھیجو، بعد میں واپس کروں گا۔", "Urdu", "Impersonation", "Scam"),
    ("بیٹا، میں امی ہوں۔ نئے فون سے میسج کر رہی ہوں۔ Rs. 12,000 بھیج دو، دوا لینی ہے۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کے بھائی کا نیا نمبر۔ فون ٹوٹ گیا ہے۔ Rs. 18,000 بھیجو، اگلے ہفتے واپس کروں گا۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کی بہن کا پیغام: بھائی مجھے بچوں کی فیس کے لیے Rs. 12,000 چاہیے۔ یہ نیا نمبر ہے۔", "Urdu", "Impersonation", "Scam"),
    ("السلام علیکم! میں آپ کا ماموں ہوں۔ بیٹے کی شادی ہے، Rs. 50,000 ادھار چاہیے۔ بعد میں واپس کروں گا۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کے آفس کا عملہ: آپ کی تنخواہ روک دی گئی ہے۔ فوری طور پر تصدیقی فیس Rs. 3,000 جمع کروائیں۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کے بچے کے اسکول کا پرنسپل: سالانہ تقریب کے لیے فوری Rs. 8,000 ادا کریں۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کا پڑوسی عمران: گھر میں ایمرجنسی ہے، بیوی کو ہسپتال لے جانا ہے۔ Rs. 10,000 ادھار چاہیے۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کی خالہ: بیٹے کا ایکسیڈنٹ ہو گیا ہے۔ Rs. 50,000 ہسپتال میں جمع کرانے ہیں۔ جلدی بھیجیں۔", "Urdu", "Impersonation", "Scam"),
    ("السلام علیکم بھائی جان۔ میں دبئی ایئرپورٹ پر ہوں۔ ویزا ختم ہو رہا ہے، Rs. 35,000 جرمانہ ادا کرنا ہے۔", "Urdu", "Impersonation", "Scam"),
    ("آپ کے چچا: دبئی سے کال کر رہا ہوں۔ پاسپورٹ کھو گیا ہے، Rs. 40,000 ایمرجنسی ٹریول دستاویز کے لیے چاہیے۔", "Urdu", "Impersonation", "Scam"),

    # ──────────────── URDU LOAN SCAMS ────────────────
    ("فوری قرضہ! صرف Rs. 2,000 پروسیسنگ فیس پر Rs. 3,00,000 تک کا لون حاصل کریں۔ کوئی ضمانت نہیں۔", "Urdu", "Loan", "Scam"),
    ("آپ کو Rs. 5,00,000 کا قرضہ پہلے سے منظور ہے۔ صرف Rs. 8,000 دستاویزی چارجز ادا کریں۔", "Urdu", "Loan", "Scam"),
    ("آپ کو بزنس لون کی پیشکش! Rs. 10,00,000 تک 0% شرح سود پر۔ Rs. 15,000 ایڈوانس فیس ادا کریں۔", "Urdu", "Loan", "Scam"),
    ("اسٹوڈنٹ لون: Rs. 10,00,000 تک تعلیمی قرضہ۔ صرف Rs. 5,000 فیس سے درخواست دیں۔", "Urdu", "Loan", "Scam"),
    ("آپ کا لون ایپلیکیشن منظور ہو گیا ہے! Rs. 5,00,000 کی رقم۔ انشورنس فیس Rs. 12,000 ادا کریں۔", "Urdu", "Loan", "Scam"),
    ("ایمرجنسی لون 1 گھنٹے میں! Rs. 50,000 سے Rs. 10,00,000 تک۔ ایڈوانس فیس صرف 2%۔", "Urdu", "Loan", "Scam"),
    ("بينا سود کا قرض چاہیے؟ اسلامک مائیکروفناننس اسکیم میں درخواست دیں۔ پروسیسنگ فیس Rs. 3,000۔", "Urdu", "Loan", "Scam"),
    ("گھر خریدنے کا خواب پورا کریں! ہوم لون Rs. 50,00,000 تک۔ ایپلیکیشن فیس Rs. 20,000۔", "Urdu", "Loan", "Scam"),

    # ──────────────── URDU FAKE CHARITY SCAMS ────────────────
    ("سیلاب متاثرین کی مدد کریں! اپنے عطیات اس نمبر پر بھیجیں۔ ہر روپیہ ضرورت مندوں تک پہنچے گا۔", "Urdu", "Fake Charity", "Scam"),
    ("یتیم بچوں کی کفالت کا ثواب کمائیں۔ ماہانہ صرف Rs. 3,000 اس نمبر پر بھیجیں۔", "Urdu", "Fake Charity", "Scam"),
    ("فلاح انسانیت فاؤنڈیشن: غریب خاندانوں کو راشن فراہم کرنے کے لیے عطیات دیں۔ بینک اکاؤنٹ نمبر: XXXX", "Urdu", "Fake Charity", "Scam"),
    ("رمضان المبارک! زکوٰۃ اور صدقہ اس نمبر پر بھیجیں۔ ہم یتیم بچوں کی کفالت کرتے ہیں۔", "Urdu", "Fake Charity", "Scam"),
    ("کشمیر زلزلہ ریلیف فنڈ: فراخدلانہ عطیات دیں۔ ایزی پیسہ نمبر: 03XX-XXXXXXX۔", "Urdu", "Fake Charity", "Scam"),
    ("عید الاضحی قربانی پروگرام: Rs. 15,000 بھیجیں اور ہم آپ کی طرف سے قربانی کریں گے۔ رسید ملے گی۔", "Urdu", "Fake Charity", "Scam"),
    ("تھر کے لیے سردیوں کے کپڑے: Rs. 3,000 عطیہ کریں۔ ٹیکس چھوٹ سرٹیفکیٹ ملے گا۔", "Urdu", "Fake Charity", "Scam"),
    ("مسجد کمیٹی: رمضان فنڈ ریزر کے لیے Rs. 5,000 اس اکاؤنٹ پر بھیجیں۔ رسید دی جائے گی۔", "Urdu", "Fake Charity", "Scam"),

    # ──────────────── MIXED LANGUAGE SCAMS (EXPANDED - diverse) ────────────────
    # Mixed: Job Scams
    ("Pakistan IT Academy ki taraf se notification: Aap ka naam international training scholarship mein select hua hai. Registration fee Rs. 5,000.", "Mixed", "Job", "Scam"),
    ("Bhai, Dubai construction company se 500 posts aayi hain. Salary AED 3000/month. Rs. 25,000 visa fee pay karein.", "Mixed", "Job", "Scam"),
    ("Work from home opportunity! Aap apne mobile se rozana Rs. 5,000 earn kar sakte hain. Bas Rs. 2,000 registration fee bhejein.", "Mixed", "Job", "Scam"),
    ("Amazon.pk warehouse job opening: 200 vacancies available. Salary Rs. 55,000/month. Pay Rs. 3,000 application fee.", "Mixed", "Job", "Scam"),
    ("Saudi Arabia mein restaurant worker chahiye. Free food aur accommodation. Agency fee Rs. 30,000 pay karein.", "Mixed", "Job", "Scam"),
    ("Freelancer.pk par 50 new jobs post hain. Aap ka profile selected hai. Premium membership Rs. 4,500 pay karein.", "Mixed", "Job", "Scam"),
    ("Qatar Airways cabin crew recruitment - salary QAR 8000/month. Rs. 35,000 medical aur training fee pay karein.", "Mixed", "Job", "Scam"),
    ("Gulf employment portal: Aap ke liye Kuwait mein 100 factory jobs available hain. Rs. 18,000 documentation fee.", "Mixed", "Job", "Scam"),
    # Mixed: Investment Scams
    ("Crypto trading bot: Rs. 10,000 invest karein aur rozana 8% profit earn karein. Guaranteed returns!", "Mixed", "Investment", "Scam"),
    ("Pakistan Stock Exchange insider tips: Rs. 50,000 invest karein, 1 week mein double hone ka mauka!", "Mixed", "Investment", "Scam"),
    ("Real estate pre-launch offer: Bahria Town Karachi mein plot Rs. 10,00,000 mein. Token money Rs. 25,000 pay karein.", "Mixed", "Investment", "Scam"),
    ("DHA Multan special scheme: Rs. 5,00,000 invest karein, 3 saal mein double value. Booking fee Rs. 20,000.", "Mixed", "Investment", "Scam"),
    ("Forex trading signals: Aap ko daily profitable trades milenge. Rs. 15,000 subscription fee pay karein.", "Mixed", "Investment", "Scam"),
    ("E-commerce business setup: Rs. 20,000 invest karein aur apna online store start karein. Rs. 1,00,000/month revenue guarantee.", "Mixed", "Investment", "Scam"),
    ("NFT marketplace: Rs. 30,000 se digital art khareedein aur 1 month mein 500% profit paayein.", "Mixed", "Investment", "Scam"),
    ("Agricultural land investment: Punjab mein 1 kanal Rs. 5,00,000. Token money Rs. 50,000 pay karein aur plot book karein.", "Mixed", "Investment", "Scam"),
    # Mixed: OTP Scams
    ("EasyPaisa se message: Aap ka account Rs. 50,000 ki suspicious activity ke liye flag hua hai. OTP share karein warna block hoga.", "Mixed", "OTP", "Scam"),
    ("JazzCash Alert: Rs. 1,00,000 ki transaction pending hai. Confirm karne ke liye OTP code share karein jo abhi mila.", "Mixed", "OTP", "Scam"),
    ("HBL Mobile banking: Sir, aap ka account hack ho raha hai. Bachane ke liye OTP code immediately bhejein.", "Mixed", "OTP", "Scam"),
    ("Meezan Bank Alert: Aap ke debit card se Rs. 75,000 ki online purchase hui hai. Block karne ke liye OTP confirm karein.", "Mixed", "OTP", "Scam"),
    ("Bank Alfalah Security: Aap ka account compromise hua hai. Recovery ke liye abhi OTP share karein.", "Mixed", "OTP", "Scam"),
    ("NADRA verification: Aap ke CNIC par multiple SIMs registered hain. OTP code share karke confirm karein warna sab band honge.", "Mixed", "OTP", "Scam"),
    # Mixed: SIM Block Scams
    ("PTA Notice: Aap ka SIM card unregistered hai. Rs. 500 pay karke registration complete karein warna 48 hours mein band hoga.", "Mixed", "SIM Block", "Scam"),
    ("Jazz: Aap ka number 24 hours mein disconnect ho jaye ga agar CNIC verification complete nahi ki. Rs. 800 pay karein.", "Mixed", "SIM Block", "Scam"),
    ("Zong Customer: Aap ke naam par 8 SIMs hain jo illegal hain. Rs. 3,000 verification fee pay karein warna police action hoga.", "Mixed", "SIM Block", "Scam"),
    ("Telenor Alert: Aap ka SIM duplicate ban raha hai kisi aur ke naam par. Rs. 1,500 pay karke rokein.", "Mixed", "SIM Block", "Scam"),
    ("Ufone: Aap ka number 7 din ke andar band ho jaye ga agar biometric verification nahi karwai. Rs. 500 fee pay karein.", "Mixed", "SIM Block", "Scam"),
    ("DIRBS Notification: Aap ka device unregistered hai. Rs. 2,000 tax pay karein warna network band ho jaye ga.", "Mixed", "SIM Block", "Scam"),
    # Mixed: Impersonation Scams
    ("Bhai, main hoon - Ammi ka naya number hai. Old phone kho gaya. Rs. 8,000 bhejo, doctor ki fee pay karni hai.", "Mixed", "Impersonation", "Scam"),
    ("Assalam o Alaikum, yeh aap ki bhabhi hai. Bhai ka phone band hai, mera naya number hai. Rs. 15,000 urgently chahiye bachon ki school fees ke liye.", "Mixed", "Impersonation", "Scam"),
    ("Beta, main Papa hoon - naye number se message kar raha hoon. Office mein emergency hai, Rs. 20,000 bhej do.", "Mixed", "Impersonation", "Scam"),
    ("Bhai jaan, main aap ka dost Ahmed hoon. Mera phone kho gaya, yeh naya number hai. Rs. 5,000 udhaar chahiye.", "Mixed", "Impersonation", "Scam"),
    ("Hello, yeh aap ki beti hai. Main college mein hoon, emergency fee pay karni hai Rs. 10,000. Papa ko mat batana.", "Mixed", "Impersonation", "Scam"),
    ("Bhai, main hoon - aap ka cousin jo Dubai mein hai. Visa issue ho gaya hai, Rs. 40,000 urgently bhejo.", "Mixed", "Impersonation", "Scam"),
    ("Ammi here, new number. Beta mujhe dawai leni hai Rs. 3,500 ki. Pharmacy walon ne udhar dene se mana kar diya.", "Mixed", "Impersonation", "Scam"),
    ("Sir, yeh aap ka tenant hai. Ghar mein pipe phat gaya hai, plumber ko Rs. 4,000 chahiye urgently.", "Mixed", "Impersonation", "Scam"),
    # Mixed: Prize/Lottery Scams
    ("Pepsi Lucky Draw 2024! Aap ka number select hua hai Rs. 25,00,000 ke liye. Processing fee Rs. 10,000 pay karein.", "Mixed", "Lottery", "Scam"),
    ("Congratulations! Aap ne Toyota Corolla jeet li hai. Delivery charges Rs. 30,000 pay karke car collect karein.", "Mixed", "Prize", "Scam"),
    ("Ramadan Mubarak Offer! Aap ne Rs. 50,00,000 ka lucky draw jeeta hai. Tax Rs. 20,000 pay karke claim karein.", "Mixed", "Lottery", "Scam"),
    ("Eid Special Prize: Aap ne umrah package win kiya hai! Visa charges Rs. 45,000 pay karein.", "Mixed", "Prize", "Scam"),
    ("Daraz.pk 11.11 Winner! Aap ko Rs. 1,00,000 ka shopping voucher mila hai. Claim fee Rs. 5,000.", "Mixed", "Prize", "Scam"),
    ("Samsung Pakistan promotion: Aap ne Galaxy Z Fold jeeta hai! Shipping fee Rs. 8,000 pay karein.", "Mixed", "Prize", "Scam"),
    ("PTV Lucky Draw! Aap ka number Rs. 10,00,000 ka winner hai. Processing charges Rs. 15,000 bhejein.", "Mixed", "Lottery", "Scam"),
    ("Shell Pakistan offer: Aap ne London trip jeeta hai! Visa aur ticket charges Rs. 50,000 pay karein.", "Mixed", "Prize", "Scam"),
    # Mixed: Bank Fraud
    ("UBL Alert: Aap ke account se London mein £500 ki transaction hui hai. Fraud hai to yeh link click karein: http://ubl-security.com", "Mixed", "Bank", "Scam"),
    ("MCB Mobile: Sir, aap ka debit card compromised hua hai. New card issue karne ke liye Rs. 1,500 courier fee pay karein.", "Mixed", "Bank", "Scam"),
    ("State Bank of Pakistan: Aap ke account ki KYC verification pending hai. CNIC details is link par enter karein: http://sbp-kyc.pk", "Mixed", "Bank", "Scam"),
    ("EasyPaisa Fraud Alert: Aap ke account se Rs. 2,00,000 transfer hue hain. Recover karne ke liye yeh link open karein.", "Mixed", "Bank", "Scam"),
    ("SadaPay Security: Aap ka account suspend hua hai. Reactivate karne ke liye Rs. 500 verification fee pay karein.", "Mixed", "Bank", "Scam"),
    ("JazzCash Notice: Aap ke naam par 3 accounts registered hain. Agar aap ne nahi khole to yeh link click karein.", "Mixed", "Bank", "Scam"),
    ("NBP Online: Aap ka account dormant ho gaya hai. Rs. 1,000 pay karke reactivate karein warna permanently band hoga.", "Mixed", "Bank", "Scam"),
    ("FBR Tax Notice: Aap par Rs. 2,00,000 ka tax outstanding hai. Is link par pay karein warna arrest warrant jaari hoga: http://fbr-tax.pk", "Mixed", "Bank", "Scam"),
    # Mixed: Loan Scams
    ("Instant Loan Approval! Aap ko Rs. 5,00,000 ka loan approved hai. Insurance fee Rs. 10,000 pay karein.", "Mixed", "Loan", "Scam"),
    ("Personal Loan without documents! Sirf CNIC par Rs. 3,00,000 milega. Processing fee Rs. 8,000.", "Mixed", "Loan", "Scam"),
    ("BISP Interest-free Loan: Aap ko Rs. 2,00,000 eligible hain. Application fee Rs. 5,000 pay karein.", "Mixed", "Loan", "Scam"),
    ("Business Loan Offer: Rs. 20,00,000 tak ke loan available hain. Sirf Rs. 15,000 documentation fee.", "Mixed", "Loan", "Scam"),
    ("Student Loan Pakistan: Rs. 10,00,000 tak educational loan. Rs. 3,000 processing fee pay karein.", "Mixed", "Loan", "Scam"),
    ("Home Renovation Loan: Rs. 8,00,000 tak qarz available hai. Application fee Rs. 12,000 pay karein.", "Mixed", "Loan", "Scam"),
]

SAFE_MESSAGES = [
    # ──────────────── LEGITIMATE JOB COMMUNICATIONS ────────────────
    ("Thank you for applying. Your interview is scheduled for Thursday at 10 AM at our Gulberg office. Please bring your CV and CNIC.", "English", "Job", "Safe"),
    ("آپ کا انٹرویو منگل کی صبح 11 بجے مقرر کیا گیا ہے۔ براہ کرم اپنے اصل دستاویزات ساتھ لائیں۔", "Urdu", "Job", "Safe"),
    ("Aap ka resume hum ne review kiya hai. Kindly office aayein Monday ko 9 bje for a formal interview. Address: Blue Area, Islamabad.", "Roman Urdu", "Job", "Safe"),
    ("We are pleased to offer you the position of Marketing Executive. Your joining date is confirmed for 1st of next month.", "English", "Job", "Safe"),
    ("Your internship certificate is ready for collection. Please visit HR department during office hours (9 AM - 5 PM).", "English", "Job", "Safe"),
    ("آپ کی ملازمت کی درخواست موصول ہو گئی ہے۔ شارٹ لسٹ ہونے پر آپ کو مطلع کیا جائے گا۔", "Urdu", "Job", "Safe"),
    ("Aap ne jo CV bheji thi, us par review ho raha hai. Next week tak inform kar dein ge.", "Roman Urdu", "Job", "Safe"),
    ("As a reminder, your probation review meeting is scheduled for Friday at 2 PM in Conference Room B.", "English", "Job", "Safe"),
    ("Office ki taraf se notice: Kal team lunch hai 1 bje. Sab employees participate karein.", "Roman Urdu", "Job", "Safe"),
    ("Human Resources: Your annual leave request for December 20-30 has been approved. Enjoy your holidays!", "English", "Job", "Safe"),
    ("نوٹس: تمام ملازمین کو مطلع کیا جاتا ہے کہ ماہانہ میٹنگ بدھ کو صبح 10 بجے ہوگی۔", "Urdu", "Job", "Safe"),
    ("Your performance appraisal form has been shared via email. Please complete and submit by end of this week.", "English", "Job", "Safe"),
    ("Aap ki salary is month 28 tareekh ko credit hogi. Bank holidays ki wajah se 1 din late ho sakti hai.", "Roman Urdu", "Job", "Safe"),
    ("Training session for new software will be held next Monday. All team members are required to attend.", "English", "Job", "Safe"),
    ("ملازمت کے معاہدے کی تجدید کے لیے براہ کرم ایچ آر ڈیپارٹمنٹ سے رابطہ کریں۔", "Urdu", "Job", "Safe"),

    # ──────────────── LEGITIMATE BANK / FINANCIAL MESSAGES ────────────────
    ("HBL: Your account ending ****4521 has been credited with Rs. 45,000 on 15-Aug. Available balance: Rs. 1,23,450.", "English", "Bank", "Safe"),
    ("آپ کے اکاؤنٹ میں Rs. 25,000 جمع ہو گئے ہیں۔ موجودہ بیلنس: Rs. 87,300۔ شکریہ۔", "Urdu", "Bank", "Safe"),
    ("Meezan Bank: Your monthly account statement for July is available. Please check your email or visit any branch.", "English", "Bank", "Safe"),
    ("Aap ki credit card bill ki due date 25 August hai. Minimum payment Rs. 5,000. JazzCash ya branch se pay karein.", "Roman Urdu", "Bank", "Safe"),
    ("UBL: ATM withdrawal of Rs. 10,000 from your account at DHA Branch ATM on 20-Aug at 3:45 PM.", "English", "Bank", "Safe"),
    ("Your Zakat deduction of Rs. 3,200 has been applied as per government regulations. Contact branch for SZR exemption.", "English", "Bank", "Safe"),
    ("آپ کا ماہانہ بینک اسٹیٹمنٹ تیار ہے۔ آن لائن بینکنگ یا موبائل ایپ سے چیک کریں۔", "Urdu", "Bank", "Safe"),
    ("Allied Bank: Your cheque book (25 leaves) is ready for collection at Model Town branch. Bring your CNIC.", "English", "Bank", "Safe"),
    ("MCB Mobile: Aap ne Rs. 15,000 transfer kiye Ahmed ko. Transaction ID: MCB20260815. Agar aap ne nahi kiya to helpline par call karein.", "Mixed", "Bank", "Safe"),
    ("Your fixed deposit of Rs. 5,00,000 will mature on 1st September. Visit your branch to renew or withdraw.", "English", "Bank", "Safe"),
    ("Bank Al Habib: Branch timing change notice - Gulshan branch will now operate from 9 AM to 4 PM effective September 1.", "English", "Bank", "Safe"),
    ("آپ کی آن لائن بینکنگ پاس ورڈ کی میعاد ختم ہو رہی ہے۔ براہ کرم برانچ میں آ کر نیا پاس ورڈ حاصل کریں۔", "Urdu", "Bank", "Safe"),
    ("Your debit card annual fee of Rs. 1,500 has been charged. For queries, call UBL helpline 021-111-825-888.", "English", "Bank", "Safe"),
    ("EasyPaisa: Rs. 8,000 successfully sent to 0312-XXXXXXX. Fee: Rs. 15. New balance: Rs. 22,400.", "Mixed", "Bank", "Safe"),
    ("آپ کے بینک اکاؤنٹ کی سالانہ زکوٰۃ کٹوتی Rs. 1,850 ہوئی ہے۔ تفصیلات برانچ سے حاصل کریں۔", "Urdu", "Bank", "Safe"),

    # ──────────────── LEGITIMATE OTP / VERIFICATION ────────────────
    ("Your OTP for JazzCash login is 847291. Do NOT share this code with anyone. Valid for 5 minutes.", "English", "OTP", "Safe"),
    ("آپ کا تصدیقی کوڈ 592831 ہے۔ یہ کوڈ کسی کے ساتھ شیئر نہ کریں۔", "Urdu", "OTP", "Safe"),
    ("EasyPaisa OTP: 384721. Use this code to complete your transaction. Never share this code with anyone, including EasyPaisa staff.", "English", "OTP", "Safe"),
    ("HBL Mobile: Your one-time password is 729184. Valid for 3 minutes. Do not share with anyone.", "English", "OTP", "Safe"),
    ("Aap ka Meezan Online login code 451923 hai. Kisi ko mat batayein. 5 minute mein expire ho jaye ga.", "Roman Urdu", "OTP", "Safe"),
    ("Daraz: Your verification code is 638291. Enter this code to reset your password. Valid for 10 minutes.", "English", "OTP", "Safe"),
    ("FoodPanda: OTP for your order confirmation is 192837. Share with rider only for delivery verification.", "English", "OTP", "Safe"),
    ("آپ کا پاس ورڈ ری سیٹ کوڈ 847392 ہے۔ یہ کوڈ کسی کو نہ بتائیں۔", "Urdu", "OTP", "Safe"),
    ("Careem: Your ride verification code is 4829. Share this with your driver to start the trip.", "English", "OTP", "Safe"),
    ("Zong prepaid: Aap ka balance recharge code 84729183 hai. Dial *111# to recharge.", "Roman Urdu", "OTP", "Safe"),
    ("Bykea: OTP for your account login is 573921. Never share this code. Bykea staff will never ask for it.", "English", "OTP", "Safe"),
    ("آپ کی ٹرانزیکشن کی تصدیق کے لیے کوڈ: 294817۔ صرف اس ویب سائٹ پر درج کریں جو آپ استعمال کر رہے ہیں۔", "Urdu", "OTP", "Safe"),

    # ──────────────── LEGITIMATE TELECOM / SIM MESSAGES ────────────────
    ("Jazz: Your package of 10GB monthly data has been activated. Validity: 30 days. Check remaining: *111#", "English", "SIM Block", "Safe"),
    ("آپ کا جاز ماہانہ پیکیج کامیابی سے فعال ہو گیا ہے۔ 30 جی بی ڈیٹا، 30 دن۔", "Urdu", "SIM Block", "Safe"),
    ("Telenor: Your balance is Rs. 342. To recharge, dial *555*CODE#. Helpline: 345.", "English", "SIM Block", "Safe"),
    ("Zong: Aap ka 4G device offer activate ho gaya hai. 5GB extra data 7 din ke liye. Enjoy karein!", "Roman Urdu", "SIM Block", "Safe"),
    ("Ufone: Your super card has been renewed. Charges: Rs. 1,200. Valid for 30 days with unlimited calls.", "English", "SIM Block", "Safe"),
    ("PTA: Your device IMEI has been verified and is compliant. No action needed.", "English", "SIM Block", "Safe"),
    ("آپ کے موبائل نمبر کی بائیو میٹرک تصدیق مکمل ہو گئی ہے۔ شکریہ۔", "Urdu", "SIM Block", "Safe"),
    ("Jazz: Aap ne 500 MB daily bundle subscribe kiya hai. Rs. 25 charges deduct ho gaye hain.", "Roman Urdu", "SIM Block", "Safe"),
    ("Telenor: Your postpaid bill for August is Rs. 2,450. Due date: 10th September. Pay via app or bank.", "English", "SIM Block", "Safe"),
    ("Zong: Sim replacement ke liye nearest franchise visit karein. Apna CNIC sath layein. Charges: Rs. 200.", "Roman Urdu", "SIM Block", "Safe"),
    ("Aap ka number portability request accept ho gaya hai. Naya SIM 48 ghante mein activate hoga.", "Roman Urdu", "SIM Block", "Safe"),
    ("Jazz World app update available. Download from Play Store for better experience. New features included.", "English", "SIM Block", "Safe"),

    # ──────────────── LEGITIMATE PROMOTIONS & GENERAL ────────────────
    ("Daraz: Your order #PK847291 has been shipped via TCS. Track at daraz.pk/track. Expected delivery: 3-5 days.", "English", "Prize", "Safe"),
    ("آپ کا آرڈر کامیابی سے بک ہو گیا ہے۔ ڈلیوری 3 سے 5 کاروباری دنوں میں ہوگی۔", "Urdu", "Prize", "Safe"),
    ("FoodPanda: Your order from Karachi Biryani has been confirmed. Estimated delivery: 35 minutes. Order #FP92847.", "English", "Prize", "Safe"),
    ("JazzCash: Rs. 50 cashback on your bill payment of Rs. 3,000. Cashback valid till 31 August.", "English", "Prize", "Safe"),
    ("Careem: Your ride from Gulberg to DHA has been completed. Fare: Rs. 450. Rate your driver on the app.", "English", "Prize", "Safe"),
    ("Aap ka InDrive offer accept ho gaya hai. Driver 5 minute mein pohanch jaye ga. Car: White Corolla, ABC-1234.", "Roman Urdu", "Prize", "Safe"),
    ("آپ کی بکنگ کی تصدیق: ہوٹل پرل کانٹی نینٹل، 2 راتیں، چیک ان 15 ستمبر۔", "Urdu", "Prize", "Safe"),
    ("EasyPaisa: Bill payment of Rs. 4,200 for K-Electric was successful. Transaction ID: EP847291.", "English", "Prize", "Safe"),
    ("Netflix: Your monthly subscription of Rs. 1,200 has been charged. Next billing date: September 26.", "English", "Prize", "Safe"),
    ("Your Amazon package was delivered today. If you didn't receive it, contact support within 48 hours.", "English", "Prize", "Safe"),
    ("Punjab Government: Vehicle token tax reminder for 2026-27. Pay online at punjab.gov.pk before October 31.", "English", "Prize", "Safe"),
    ("Aap ne jo mobile online order kiya tha, woh dispatch ho gaya hai. Tracking ID email pe bhej diya hai.", "Roman Urdu", "Prize", "Safe"),
    ("Pakistan Post: Your parcel has arrived at GPO Lahore. Collect within 7 days with your CNIC.", "English", "Prize", "Safe"),
    ("WAPDA bill reminder: Aap ka bijli ka bill Rs. 8,500 hai. Due date 20 September. Late surcharge apply hoga.", "Mixed", "Prize", "Safe"),
    ("Sui Gas: Your monthly bill is Rs. 2,100. Pay before 15th to avoid disconnection. Helpline: 1199.", "English", "Prize", "Safe"),

    # ──────────────── LEGITIMATE INVESTMENT / FINANCIAL ADVICE ────────────────
    ("PSX Monthly Review: KSE-100 index gained 2.3% in August. Consult your licensed broker for investment decisions.", "English", "Investment", "Safe"),
    ("آپ کے میوچل فنڈ کی ماہانہ کارکردگی رپورٹ: اگست میں 1.8% ریٹرن۔ تفصیلات ای میل میں بھیجی گئی ہیں۔", "Urdu", "Investment", "Safe"),
    ("National Savings: Your Behbood Certificate profit of Rs. 12,500 has been credited to your account for August.", "English", "Investment", "Safe"),
    ("Aap ki PLS account ki monthly profit Rs. 3,400 credit ho gayi hai. Statement bank app mein check karein.", "Roman Urdu", "Investment", "Safe"),
    ("SECP Advisory: Always verify investment companies on secp.gov.pk before investing. Report fraud to 051-9207078.", "English", "Investment", "Safe"),
    ("Your Roshan Digital Account statement for Q2 2026 is available on the SBP portal. Login to view details.", "English", "Investment", "Safe"),
    ("آپ کے اسٹاک پورٹ فولیو کی ہفتہ وار اپ ڈیٹ: مجموعی ریٹرن +3.2%۔ تفصیلات ای میل میں دیکھیں۔", "Urdu", "Investment", "Safe"),
    ("Al Meezan Mutual Fund: NAV per unit as of August 25 is Rs. 142.35. Past performance is not indicative of future returns.", "English", "Investment", "Safe"),
    ("Aap ka SIP installment of Rs. 5,000 successfully invest ho gaya hai AKD Mutual Fund mein. Units allotted: 35.2", "Roman Urdu", "Investment", "Safe"),
    ("FBR: Your income tax return for FY 2025-26 has been acknowledged. Refund processing is underway.", "English", "Investment", "Safe"),

    # ──────────────── LEGITIMATE CHARITY / SOCIAL ────────────────
    ("Edhi Foundation: Thank you for your donation of Rs. 5,000. Receipt #ED-847291. Tax deductible under section 2(36).", "English", "Fake Charity", "Safe"),
    ("Shaukat Khanum: Your monthly donation of Rs. 2,000 has been received via auto-debit. JazakAllah!", "Mixed", "Fake Charity", "Safe"),
    ("آپ کا عطیہ وصول ہو گیا ہے۔ رسید نمبر: SK-48291۔ آپ کا تعاون قابل قدر ہے۔", "Urdu", "Fake Charity", "Safe"),
    ("Saylani Welfare: Aap ki ration bag donation ki receipt email par bhej di gayi hai. Thank you for your support.", "Roman Urdu", "Fake Charity", "Safe"),
    ("Chhipa: Your Eid Qurbani booking is confirmed. You will receive meat distribution photos via SMS. Thank you.", "English", "Fake Charity", "Safe"),
    ("Aap ne Indus Hospital ko Rs. 10,000 donate kiye. Tax exemption certificate 30 din mein email hoga.", "Roman Urdu", "Fake Charity", "Safe"),
    ("LRBT: Your free eye surgery camp registration is confirmed for September 5 at LRBT Hospital, Karachi.", "English", "Fake Charity", "Safe"),
    ("آغا خان فاؤنڈیشن: آپ کی ماہانہ ڈونیشن Rs. 3,000 کامیابی سے ڈیبٹ ہو گئی ہے۔", "Urdu", "Fake Charity", "Safe"),
    ("Al-Khidmat Foundation: Your orphan child sponsorship for August has been processed. Progress report will be shared quarterly.", "English", "Fake Charity", "Safe"),
    ("Shifa Foundation: Aap ki Sadqa-e-Jariya water well project mein Rs. 5,000 ki donation mil gayi hai. JazakAllah Khair.", "Mixed", "Fake Charity", "Safe"),

    # ──────────────── LEGITIMATE PERSONAL / FAMILY ────────────────
    ("Mom: Beta dinner ke liye jaldi aa jana. Aaj biryani bani hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Assalam o Alaikum bhai, kal Eid milne aayein ge. Sab family members ko salam kehna.", "Roman Urdu", "Impersonation", "Safe"),
    ("Papa: Ghar pohanch gaya hoon. Sab theek hai. Kal shaam tak wapis aaunga.", "Roman Urdu", "Impersonation", "Safe"),
    ("آپ کی اماں کا پیغام: بیٹا دوا لینا مت بھولنا۔ شام کو فون کرنا۔", "Urdu", "Impersonation", "Safe"),
    ("Bhai, kal cricket match hai shaam 5 bje ground mein. Bat aur gloves le aana.", "Roman Urdu", "Impersonation", "Safe"),
    ("Sis: Can you pick up kids from school today? I have a dentist appointment at 3 PM.", "English", "Impersonation", "Safe"),
    ("Chacha: Gaon mein sab khairiyat hai. Ammi ki tabiyat behtar hai. Dua karein.", "Roman Urdu", "Impersonation", "Safe"),
    ("بیٹی، کل ہم تمہارے گھر آ رہے ہیں۔ دوپہر کا کھانا تیار رکھنا۔", "Urdu", "Impersonation", "Safe"),
    ("Bhai, aaj office se jaldi nikal jana. Traffic bohat hogi kyunke road closure hai I.I Chundrigar pe.", "Roman Urdu", "Impersonation", "Safe"),
    ("Uncle: Aap ki car ki service ho gayi hai. Workshop se le jayein. Bill Rs. 12,000.", "Roman Urdu", "Impersonation", "Safe"),
    ("Dear, groceries list bana lo. Doodh, dahi, roti, aur sabzi chahiye. Main 7 bje tak aaungi.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhaijaan: Walima ki date fix ho gayi hai - 15 October. Venue: Pearl Continental. Dawat card bhej raha hoon.", "Roman Urdu", "Impersonation", "Safe"),

    # ──────────────── LEGITIMATE LOAN / CREDIT ────────────────
    ("HBL: Your personal loan of Rs. 5,00,000 has been approved. Monthly installment: Rs. 12,500. Tenure: 5 years.", "English", "Loan", "Safe"),
    ("آپ کی قرضہ درخواست منظور ہو گئی ہے۔ ماہانہ قسط: Rs. 8,000۔ تفصیلات برانچ سے حاصل کریں۔", "Urdu", "Loan", "Safe"),
    ("Meezan Bank: Your car financing Ijarah agreement is ready. Please visit DHA branch to sign documents.", "English", "Loan", "Safe"),
    ("Aap ka credit card statement August ka available hai. Total outstanding: Rs. 35,000. Minimum payment: Rs. 5,000.", "Roman Urdu", "Loan", "Safe"),
    ("Kissan Bank: Agricultural loan of Rs. 2,00,000 sanctioned. Visit your nearest branch with CNIC and land documents.", "English", "Loan", "Safe"),
    ("Student loan update: Your HEC loan application status is 'Under Review'. Check at hec.gov.pk.", "English", "Loan", "Safe"),
    ("آپ کے گھر کے قرضے کی ماہانہ قسط Rs. 25,000 کل ڈیبٹ ہو جائے گی۔", "Urdu", "Loan", "Safe"),
    ("Microfinance Bank: Aap ka chota karobar loan Rs. 1,00,000 approve hua hai. Disbursement 3 working days mein.", "Roman Urdu", "Loan", "Safe"),
    ("Your loan EMI of Rs. 18,500 will be auto-debited on 5th of every month. Ensure sufficient balance.", "English", "Loan", "Safe"),
    ("BankIslami: Diminishing Musharakah home finance pre-approval letter is ready. Collect from F-7 branch.", "English", "Loan", "Safe"),

    # ──────────────── ADDITIONAL SAFE MESSAGES ────────────────
    ("Your Zong monthly bill of Rs. 1,850 has been generated. Due date: 5th September. Pay via app or jazzcash.", "English", "SIM Block", "Safe"),
    ("Aap ki Nayapay account ki monthly statement available hai. App mein check karein.", "Roman Urdu", "Bank", "Safe"),
    ("SadaPay: Rs. 500 sent to Ali Hassan. Reference: dinner split. New balance: Rs. 12,430.", "English", "Bank", "Safe"),
    ("Govt of Punjab: Your driving license renewal fee of Rs. 2,000 has been received. Visit BRTA office for biometric.", "English", "Prize", "Safe"),
    ("آپ کی پاسپورٹ کی درخواست جمع ہو گئی ہے۔ پروسیسنگ میں 10 کاروباری دن لگیں گے۔", "Urdu", "Prize", "Safe"),
    ("Remittance received: Rs. 1,50,000 from UAE via Western Union credited to your MCB account.", "English", "Bank", "Safe"),
    ("NADRA: Your CNIC renewal application has been submitted. Collection date: 15 September at Saddar office.", "English", "Prize", "Safe"),
    ("Bhai, main bazaar ja raha hoon. Kuch chahiye to batao. Doodh, roti?", "Roman Urdu", "Impersonation", "Safe"),
    ("School notice: Annual exam schedule has been updated. Check school website for date sheet.", "English", "Prize", "Safe"),
    ("آپ کے بچے کی اسکول فیس Rs. 15,000 موصول ہو گئی ہے۔ رسید محفوظ رکھیں۔", "Urdu", "Prize", "Safe"),
    ("Bykea: Your ride fare was Rs. 280 from Saddar to Clifton. Thank you for riding with Bykea!", "English", "Prize", "Safe"),
    ("Aap ki Sehat Card registration confirm ho gayi hai. Hospital mein free treatment milega.", "Roman Urdu", "Prize", "Safe"),
    ("Ehsaas Program: Aap ki eligibility check complete. Result available on 8171 web portal. Visit to check.", "English", "Prize", "Safe"),
    ("PakSim.info: Vehicle registration check for ABC-1234 completed. Owner: Ahmed Khan, Lahore.", "English", "Prize", "Safe"),
    ("آپ کا ہیلتھ انشورنس کلیم منظور ہو گیا ہے۔ رقم 5 کاروباری دنوں میں اکاؤنٹ میں آ جائے گی۔", "Urdu", "Prize", "Safe"),
    ("Auntie: Beta kal shaam ko chai pe aana. Ghar mein naye mehmaan aa rahe hain.", "Roman Urdu", "Impersonation", "Safe"),
    ("Your K-Electric bill payment of Rs. 7,800 was successful via JazzCash. Transaction ref: JC847291.", "English", "Prize", "Safe"),
    ("Office IT: Your laptop has been scheduled for annual maintenance. Please submit to IT by Friday.", "English", "Job", "Safe"),
    ("Beta, ghar ka rashan khatam ho raha hai. Aate waqt aata, cheeni aur chai le aana.", "Roman Urdu", "Impersonation", "Safe"),
    ("Pakistan Post: Your registered letter has been delivered.签收 confirmation sent to your address.", "English", "Prize", "Safe"),
    ("University notice: Fall semester classes begin September 1. Fee submission deadline: August 30.", "English", "Prize", "Safe"),
    ("Aap ka SSGC gas meter reading submit ho gayi hai. Bill 3 din mein generate hoga.", "Roman Urdu", "Prize", "Safe"),
    ("Indigo Hospital: Aap ka appointment Dr. Ahmed ke saath kal 11 bje confirm hai. OPD card sath layein.", "Roman Urdu", "Prize", "Safe"),
    ("آپ کا ووٹر رجسٹریشن کارڈ تیار ہے۔ قریبی نادرا دفتر سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("Baji, mehndi ki shopping kal karein ge. Main list bana leti hoon aaj raat tak.", "Roman Urdu", "Impersonation", "Safe"),
    ("FBR: Your NTN registration certificate has been issued. View on iris.fbr.gov.pk portal.", "English", "Investment", "Safe"),
    ("Aap ka pension payment order (PPO) ready hai. District accounts office se collect karein.", "Roman Urdu", "Loan", "Safe"),
    ("Metro Cash & Carry: Your online order #MC-92847 is packed and ready for pickup at counter 3.", "English", "Prize", "Safe"),
    ("آپ کے بچے کا ویکسینیشن شیڈول: اگلی ویکسین 10 ستمبر کو۔ براہ کرم ہسپتال تشریف لائیں۔", "Urdu", "Prize", "Safe"),
    ("Bhai, kal subah 6 bje flight hai. Taxi 4:30 pe aa jayegi. Sab documents ready rakhna.", "Roman Urdu", "Impersonation", "Safe"),

    # ──────────────── LEGITIMATE OTP / VERIFICATION (EXPANDED - anti-FP) ────────────────
    ("InDrive: Your ride code is 3847. Share this code with the driver to verify the trip.", "English", "OTP", "Safe"),
    ("Yango: Your trip verification code is 7291. Show this to your driver before starting.", "English", "OTP", "Safe"),
    ("Swvl: OTP for your booking confirmation is 491827. Valid for 10 minutes. Do not share with anyone.", "English", "OTP", "Safe"),
    ("Aap ka SadaPay login code 829174 hai. Yeh code kisi ko mat batayein. 5 minute tak valid hai.", "Roman Urdu", "OTP", "Safe"),
    ("Nayapay: Your transaction verification code is 573912. Never share this code with anyone, including Nayapay staff.", "English", "OTP", "Safe"),
    ("Aap ka NADRA portal login code 394817 hai. Kisi ko share na karein. 10 minute valid.", "Roman Urdu", "OTP", "Safe"),
    ("Yayvo: Your order verification code is 847291. Enter this code to confirm your purchase.", "English", "OTP", "Safe"),
    ("BISP: Aap ka eligibility check code 583921 hai. Is code ko 8171 par bhejein. Kisi aur ko na batayein.", "Roman Urdu", "OTP", "Safe"),
    ("آپ کا نادرا آن لائن پورٹل کا تصدیقی کوڈ 729184 ہے۔ یہ کوڈ کسی کو نہ بتائیں۔", "Urdu", "OTP", "Safe"),
    ("Ehsaas Program: Your verification code is 384721. Enter on the 8171 portal to check eligibility. Do not share.", "English", "OTP", "Safe"),
    ("FBR: Your tax filing OTP is 648291. Enter this code on iris.fbr.gov.pk to complete your return.", "English", "OTP", "Safe"),
    ("Telemart: Your order confirmation code is 573921. Use this to track your delivery.", "English", "OTP", "Safe"),
    ("Aap ka JazzCash mobile app login code 847192 hai. Kisi ko share na karein. JazzCash staff kabhi code nahi mangta.", "Roman Urdu", "OTP", "Safe"),
    ("Roshan Digital Account: Your login OTP is 392847. Valid for 5 minutes. Never share with anyone.", "English", "OTP", "Safe"),
    ("آپ کا ایزی پیسہ ایپ کا لاگ ان کوڈ 582917 ہے۔ یہ کوڈ کسی کے ساتھ شیئر نہ کریں۔", "Urdu", "OTP", "Safe"),
    ("HBL Mobile: Your fund transfer OTP is 748291. Do NOT share this code with anyone. Valid for 3 minutes.", "English", "OTP", "Safe"),
    ("Meezan Mobile: Your transaction authorization code is 928374. Never share with anyone including bank staff.", "English", "OTP", "Safe"),
    ("UBL Digital: Aap ka one-time password 473921 hai. Kisi ko mat batayein. 5 minute mein expire ho jaye ga.", "Roman Urdu", "OTP", "Safe"),
    ("Bank Alfalah: Your bill payment OTP is 582917. Enter this code to confirm. Do not share with anyone.", "English", "OTP", "Safe"),
    ("Careem: Aap ka ride verification code 4829 hai. Yeh code driver ko dikhayein. Kisi aur ko na batayein.", "Roman Urdu", "OTP", "Safe"),
    ("Bykea: Your account login OTP is 739281. Never share this code. Bykea staff will never ask for it.", "English", "OTP", "Safe"),
    ("FoodPanda: Your payment verification code is 483921. Do not share with anyone.", "English", "OTP", "Safe"),
    ("آپ کا میزان بینک آن لائن تصدیقی کوڈ 847291 ہے۔ بینک عملہ کبھی کوڈ نہیں مانگتا۔", "Urdu", "OTP", "Safe"),
    ("Zong prepaid: Aap ka recharge code 84729183 hai. Dial *111# to recharge. Keep this code private.", "Roman Urdu", "OTP", "Safe"),
    ("Daraz: Your password reset code is 592837. Enter this code to change your password. Valid 10 minutes.", "English", "OTP", "Safe"),

    # ──────────────── TELECOM SERVICE MESSAGES (EXPANDED - anti-FP) ────────────────
    ("Jazz: Your Monthly Super bundle has been activated successfully. 30GB data + 3000 mins. Valid 30 days. Charges: Rs. 999.", "English", "SIM Block", "Safe"),
    ("Zong: Aap ka Weekly Internet package activate ho gaya hai. 10GB data, 7 din valid. Rs. 150 charges deduct hue hain.", "Roman Urdu", "SIM Block", "Safe"),
    ("Telenor: Your 3G/4G monthly package has been renewed. 50GB data. Charges: Rs. 1,500. Validity: 30 days.", "English", "SIM Block", "Safe"),
    ("Ufone: Your Super Card has been auto-renewed. Monthly charges Rs. 1,199 deducted. Unlimited calls + 15GB data active.", "English", "SIM Block", "Safe"),
    ("Jazz: Aap ka data bundle expire ho raha hai kal. Renew karne ke liye *117# dial karein. New packages available.", "Roman Urdu", "SIM Block", "Safe"),
    ("Zong: Your prepaid balance is Rs. 847. To recharge, dial *222*CODE#. Helpline: 310.", "English", "SIM Block", "Safe"),
    ("Telenor: Aap ka postpaid bill Rs. 3,200 generate ho gaya hai. Due date: 15 September. Pay via app or Easypaisa.", "Roman Urdu", "SIM Block", "Safe"),
    ("Jazz World: Aap ne 2GB daily bundle subscribe kiya hai. Rs. 35 charges deduct ho gaye. Valid: 1 day.", "Roman Urdu", "SIM Block", "Safe"),
    ("Ufone: Your international roaming package has been activated. 500MB data valid for 7 days in UAE. Charges: Rs. 2,500.", "English", "SIM Block", "Safe"),
    ("PTA: Your device IMEI 354XXXXX has been registered and verified. Your phone is compliant with PTA regulations.", "English", "SIM Block", "Safe"),
    ("Zong: Aap ka 4G SIM upgrade complete ho gaya hai. Enjoy faster internet speeds. No extra charges.", "Roman Urdu", "SIM Block", "Safe"),
    ("Telenor: Your caller tune 'Mere Rashke Qamar' has been activated. Monthly charges: Rs. 30. Unsubscribe: *230#.", "English", "SIM Block", "Safe"),
    ("Jazz: Aap ka number port ho gaya hai Telenor se. Naya SIM 24 ghante mein activate hoga. Welcome to Jazz!", "Roman Urdu", "SIM Block", "Safe"),
    ("Ufone: Your voice mail service has been activated. Dial 3333 to check messages. Monthly charges: Rs. 50.", "English", "SIM Block", "Safe"),
    ("Zong: Aap ka monthly bill Rs. 2,450 generate ho gaya hai. Due date: 5 October. Pay via JazzCash ya bank.", "Roman Urdu", "SIM Block", "Safe"),
    ("Telenor: Your call package of 5000 mins has been activated. Valid for 30 days. Rs. 450 deducted.", "English", "SIM Block", "Safe"),
    ("Jazz: Aap ka SIM replacement complete ho gaya hai. Naya SIM 2 ghante mein activate hoga. Charges: Rs. 150.", "Roman Urdu", "SIM Block", "Safe"),
    ("Ufone: Your balance save plan is active. Rs. 500 bonus credit added. Valid for 15 days on calls only.", "English", "SIM Block", "Safe"),
    ("PTA: DIRBS verification successful. Your device is compliant and will continue to work on all Pakistani networks.", "English", "SIM Block", "Safe"),
    ("Zong: Aap ka postpaid plan auto-renew ho gaya hai. Monthly charges Rs. 1,800. 20GB data + unlimited calls.", "Roman Urdu", "SIM Block", "Safe"),
    ("Telenor: Your Easypaisa auto-pay for monthly bill has been set up successfully. Rs. 2,100 will deduct on 1st of each month.", "English", "SIM Block", "Safe"),
    ("Jazz: Data sharing activated. Share up to 5GB with 2 family numbers. Monthly charges: Rs. 100.", "English", "SIM Block", "Safe"),
    ("آپ کا جاز سم کارڈ بائیو میٹرک تصدیق کے بعد کامیابی سے بحال ہو گیا ہے۔ شکریہ۔", "Urdu", "SIM Block", "Safe"),
    ("Ufone: Your super card charges Rs. 1,200 successfully deducted. Valid for 30 days with unlimited calls and 10GB.", "English", "SIM Block", "Safe"),
    ("Zong: Your international calling bundle activated. 200 mins to USA/UK. Charges Rs. 500. Valid 30 days.", "English", "SIM Block", "Safe"),

    # ──────────────── SERVICE NOTIFICATIONS (EXPANDED - anti-FP) ────────────────
    ("Careem: Aap ki ride Gulberg se DHA tak complete ho gayi hai. Fare: Rs. 380. Payment: Cash. Rate your driver.", "Roman Urdu", "Prize", "Safe"),
    ("InDrive: Your ride from Model Town to Airport is confirmed. Agreed fare: Rs. 650. Driver arriving in 5 mins.", "English", "Prize", "Safe"),
    ("Bykea: Aap ki ride Saddar se Clifton tak complete. Fare: Rs. 220. Thank you for riding with Bykea!", "Roman Urdu", "Prize", "Safe"),
    ("Careem: Aap ki ride cancel ho gayi hai. Cancellation charges: Rs. 50 deducted from your wallet.", "Roman Urdu", "Prize", "Safe"),
    ("Yango: Your trip from Bahria Town to F-8 completed. Fare: Rs. 490. Rate your experience on the app.", "English", "Prize", "Safe"),
    ("InDrive: Driver accepted your offer of Rs. 350 for trip from Johar Town to DHA. Car: White City, LE-5821.", "English", "Prize", "Safe"),
    ("Careem: Your wallet has been topped up with Rs. 1,000 via JazzCash. Current balance: Rs. 1,350.", "English", "Prize", "Safe"),
    ("PakWheels: Your vehicle verification report for ABC-1234 is ready. Owner: Ahmed Khan. Report sent to email.", "English", "Prize", "Safe"),
    ("Excise & Taxation Punjab: Your vehicle token tax of Rs. 3,500 for 2026-27 has been received. Receipt #ET-847291.", "English", "Prize", "Safe"),
    ("K-Electric: Aap ka bijli ka bill Rs. 12,500 generate ho gaya hai. Due date: 20 October. Pay via JazzCash ya bank.", "Roman Urdu", "Prize", "Safe"),
    ("SSGC: Your gas bill for September is Rs. 3,200. Due date: 15 October. Pay online at ssgc.com.pk.", "English", "Prize", "Safe"),
    ("LESCO: Aap ka bijli ka bill Rs. 8,900 ready hai. Due date: 25 September. Late payment surcharge: 10%.", "Roman Urdu", "Prize", "Safe"),
    ("SNGPL: Gas bill reminder - Rs. 4,500 due by October 10. Pay via Easypaisa, JazzCash, or bank branch.", "English", "Prize", "Safe"),
    ("PTCL: Your broadband bill Rs. 2,800 has been generated. Due date: 5 October. Pay via app or bank.", "English", "Prize", "Safe"),
    ("Daraz: Your order #PK928471 has been delivered successfully. Please confirm receipt in the app. Thank you for shopping!", "English", "Prize", "Safe"),
    ("Telemart: Aap ka order #TM-48291 dispatch ho gaya hai via Leopards. Tracking ID sent to email.", "Roman Urdu", "Prize", "Safe"),
    ("FoodPanda: Aap ka order Karachi Biryani se confirm ho gaya hai. Estimated delivery: 30 minutes. Order #FP39281.", "Roman Urdu", "Prize", "Safe"),
    ("Cheetay: Your parcel has been picked up and is on its way. Track at cheetay.com/track/TRK-847291.", "English", "Prize", "Safe"),
    ("Aga Khan Hospital: Aap ka appointment Dr. Fatima ke saath kal subah 10 bje confirm hai. OPD card sath layein.", "Roman Urdu", "Prize", "Safe"),
    ("Shifa International: Your lab test results are ready. Login to patient portal to view reports.", "English", "Prize", "Safe"),
    ("Chughtai Lab: Aap ki blood test report email par bhej di gayi hai. Doctor se consult karein.", "Roman Urdu", "Prize", "Safe"),
    ("Indus Hospital: Your appointment for dental checkup is confirmed on 15 Oct at 2 PM. Free consultation.", "English", "Prize", "Safe"),
    ("LUMS: Fall semester fee of Rs. 2,85,000 is due by September 15. Pay via HBL or online banking.", "English", "Prize", "Safe"),
    ("NUST: Your exam date sheet for Fall 2026 has been uploaded. Check LMS portal for details.", "English", "Prize", "Safe"),
    ("آپ کے بچے کی اسکول کی سالانہ تقریب 20 اکتوبر کو ہوگی۔ والدین سے شرکت کی درخواست ہے۔", "Urdu", "Prize", "Safe"),
    ("PakSim.info: Vehicle ownership transfer for XYZ-5678 completed. New owner: Ali Raza, Karachi.", "English", "Prize", "Safe"),
    ("NADRA: Aap ka CNIC renewal process complete ho gaya hai. Naya card 15 din mein aap ke address par bhej diya jaye ga.", "Roman Urdu", "Prize", "Safe"),
    ("EasyPaisa: Bill payment of Rs. 6,500 for PTCL was successful. Transaction ID: EP928471. Receipt saved in app.", "English", "Prize", "Safe"),
    ("JazzCash: Aap ne Rs. 15,000 successfully send kiye Ahmed ko. Transaction ID: JC847291. New balance: Rs. 8,400.", "Roman Urdu", "Prize", "Safe"),
    ("Netflix: Aap ki monthly subscription Rs. 1,500 charge ho gayi hai. Next billing: October 26.", "Roman Urdu", "Prize", "Safe"),
    ("Spotify: Your Premium subscription renewed. Rs. 299 charged. Enjoy ad-free music for 30 days.", "English", "Prize", "Safe"),
    ("Careem: Aap ki ride complete ho gayi hai. Fare Rs. 520 wallet se deduct hua. Driver rating: 4.8 stars.", "Roman Urdu", "Prize", "Safe"),
    ("InDrive: Your cargo delivery from Gulberg to Cantt completed. Fare: Rs. 800. Thank you!", "English", "Prize", "Safe"),
    ("Bykea: Aap ka parcel delivery complete. Fare: Rs. 350. Rider: Ahmed. Rate your experience.", "Roman Urdu", "Prize", "Safe"),
    ("WAPDA: Aap ka electricity bill Rs. 15,200 generate ho gaya hai. Due date: 25 October. Pay before deadline.", "Roman Urdu", "Prize", "Safe"),
    ("Sui Northern Gas: Your meter reading has been submitted. Bill will generate in 3 working days.", "English", "Prize", "Safe"),
    ("آپ کا ہیلتھ انشورنس پریمیم Rs. 5,000 موصول ہو گیا ہے۔ کوریج ایک سال کے لیے فعال ہے۔", "Urdu", "Prize", "Safe"),
    ("Sehat Card: Aap ki family coverage confirm ho gayi hai. Free treatment at all panel hospitals.", "Roman Urdu", "Prize", "Safe"),
    ("BISP: Aap ki quarterly payment Rs. 9,000 ready hai. Nazdeeki BISP center se CNIC ke saath collect karein.", "Roman Urdu", "Prize", "Safe"),
    ("Ehsaas Kafalat: Aap ki eligibility confirm ho gayi hai. Next payment: October 2026. Visit nearest HBL branch.", "Roman Urdu", "Prize", "Safe"),

    # ──────────────── LEGITIMATE PERSONAL/FAMILY (EXPANDED) ────────────────
    ("Ammi: Beta aaj khana mat banana, main biryani bana rahi hoon. Shaam tak aa jana.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai, cricket match dekhne chalein? Pakistan vs India aaj raat 8 bje. Snacks main laata hoon.", "Roman Urdu", "Impersonation", "Safe"),
    ("Papa: Office se nikal gaya hoon. 30 minute mein ghar pohanch jaonga. Darwaza kholna.", "Roman Urdu", "Impersonation", "Safe"),
    ("Beta, kal chutti hai na? Ghar aa jana, kuch zaroori kaam hai. Dadi bhi aayi hain.", "Roman Urdu", "Impersonation", "Safe"),
    ("Sis: Mera laptop kharab ho gaya. Kal office se IT wale ko dikha dena please.", "Roman Urdu", "Impersonation", "Safe"),
    ("Uncle: Kal walima hai, shaam 7 bje Pearl Continental. Aap sab family ko lekar aayein.", "Roman Urdu", "Impersonation", "Safe"),
    ("Aunty: Beta aaj sham ko ghar aana, kheer banai hai. Aur haan, apni sister ko bhi le aana.", "Roman Urdu", "Impersonation", "Safe"),
    ("Mom: Fridge mein khaana rakha hai. Garam karke kha lena. Main 9 bje tak aaungi.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai: Kal subah gym chalein? 6 bje milte hain. Trainer ne naya plan banaya hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Chacha: Gaon se sab khairiyat se hain. Fasal achhi hui hai is saal. Dua karein.", "Roman Urdu", "Impersonation", "Safe"),
    ("Dear: Grocery list ready karo. Main office se aate waqt le aaungi.", "Roman Urdu", "Impersonation", "Safe"),
    ("Papa: Kal school ki PTM hai, 11 bje. Main chala jaonga. Tum chutti kar lena.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhaijaan: Eidi tayyar hai. Eid ki subah milne aana. Sab family ko Eidi milegi.", "Roman Urdu", "Impersonation", "Safe"),
    ("Cousin: Aaj raat dinner plan hai DHA mein. 8 bje milte hain at Xanders restaurant.", "Roman Urdu", "Impersonation", "Safe"),
    ("Office buddy: Kal presentation hai, slides ready kar lena. Boss ko 10 bje dikhani hain.", "Roman Urdu", "Impersonation", "Safe"),
    ("Baji: Kal mehndi hai, subah 10 bje aana. Shopping complete ho gayi hai kal ki.", "Roman Urdu", "Impersonation", "Safe"),
    ("Mom: Beta ghar aate waqt doodh aur bread le aana. Fridge mein khatam ho gaya hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai, wifi ka bill pay kar diya? Internet slow chal raha hai aaj kal.", "Roman Urdu", "Impersonation", "Safe"),
    ("Beti, weekend pe aa jana. Tumhare favourite pakode banaungi. Dadi bhi miss kar rahi hain.", "Roman Urdu", "Impersonation", "Safe"),
    ("Friend: Yar kal ka plan confirm karo. Movie dekhne chalein ya cricket khelein?", "Roman Urdu", "Impersonation", "Safe"),
    ("آپ کے ابو کا پیغام: بیٹا شام کو دفتر سے آتے ہوئے سبزیاں لیتے آنا۔", "Urdu", "Impersonation", "Safe"),
    ("Bhai: Office party hai Friday ko. Sab colleagues aa rahe hain. Venue final nahi hua abhi.", "Roman Urdu", "Impersonation", "Safe"),
    ("Sis: Baby shower ki planning kar rahe hain. Next Sunday ko ghar pe. Sab arrangements karni hain.", "Roman Urdu", "Impersonation", "Safe"),
    ("Papa: Car ki servicing ho gayi hai. Workshop se le aana. Bill paid already.", "Roman Urdu", "Impersonation", "Safe"),
    ("Ammi: Kal nani aa rahi hain. Ghar saaf kar dena aur AC chala dena unke kamre mein.", "Roman Urdu", "Impersonation", "Safe"),

    # ──────────────── LEGITIMATE BANK/FINANCIAL (EXPANDED) ────────────────
    ("HBL: Rs. 75,000 credited to your account ****8821 from ABC Company (Salary Aug 2026). Available balance: Rs. 1,98,450.", "English", "Bank", "Safe"),
    ("Meezan Bank: Your auto-debit for Rs. 12,500 (home loan EMI) was successful on 5 Aug. Next debit: 5 Sep.", "English", "Bank", "Safe"),
    ("Aap ki SadaPay se Rs. 3,500 ki payment successful rahi Ahmed ko. New balance: Rs. 18,900.", "Roman Urdu", "Bank", "Safe"),
    ("JazzCash: Bill payment Rs. 7,800 K-Electric successful. Transaction ID: JC928471. Receipt saved in app.", "English", "Bank", "Safe"),
    ("EasyPaisa: Rs. 25,000 received from 0312-XXXXXXX. New balance: Rs. 45,600. Transaction confirmed.", "English", "Bank", "Safe"),
    ("Nayapay: Your card transaction of Rs. 1,250 at Starbucks DHA was successful. Available balance: Rs. 12,350.", "English", "Bank", "Safe"),
    ("UBL: Your salary of Rs. 1,50,000 has been credited. Account ending ****4521. Date: 28 Aug.", "English", "Bank", "Safe"),
    ("Aap ka Meezan Bank monthly statement August ka email par bhej diya gaya hai. Check karein.", "Roman Urdu", "Bank", "Safe"),
    ("MCB: Your credit card bill of Rs. 28,500 has been generated. Minimum payment Rs. 5,700. Due: 15 Sep.", "English", "Bank", "Safe"),
    ("Bank Alfalah: Aap ka auto-debit Rs. 8,500 (car loan installment) successful. Next payment: 1 Oct.", "Roman Urdu", "Bank", "Safe"),

    # ──────────────── URDU BANK/FINANCIAL SAFE (anti-FP) ────────────────
    ("آپ کے بینک اکاؤنٹ میں ماہانہ منافع Rs. 3,500 جمع ہو گیا ہے۔ تفصیلات برانچ سے حاصل کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کے اکاؤنٹ سے ماہانہ سروس چارجز Rs. 250 کٹوتی ہوئی ہے۔ شکریہ۔", "Urdu", "Bank", "Safe"),
    ("آپ کا بینک اکاؤنٹ اسٹیٹمنٹ تیار ہے۔ آن لائن بینکنگ یا موبائل ایپ سے چیک کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کی زکوٰۃ کٹوتی خود بخود ہو گئی ہے۔ تفصیلات برانچ سے معلوم کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کے اکاؤنٹ میں Rs. 50,000 جمع ہو گئے ہیں۔ موجودہ بیلنس: Rs. 2,45,000۔", "Urdu", "Bank", "Safe"),
    ("آپ کا ماہانہ بینک اسٹیٹمنٹ ای میل پر بھیج دیا گیا ہے۔ برائے مہربانی چیک کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کی چیک بک تیار ہے۔ برانچ سے شناختی کارڈ کے ساتھ حاصل کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کے فکسڈ ڈپازٹ کی میعاد پوری ہو گئی ہے۔ تجدید کے لیے برانچ تشریف لائیں۔", "Urdu", "Bank", "Safe"),
    ("آپ کے ڈیبٹ کارڈ کی سالانہ فیس Rs. 1,500 چارج ہو گئی ہے۔", "Urdu", "Bank", "Safe"),
    ("آپ کی آن لائن بینکنگ کی ماہانہ حد Rs. 5,00,000 ہے۔ تفصیلات ہیلپ لائن سے حاصل کریں۔", "Urdu", "Bank", "Safe"),

    # ──────────────── URDU GOVERNMENT/NADRA SAFE (anti-FP) ────────────────
    ("آپ کا شناختی کارڈ کی تجدید کی درخواست موصول ہو گئی ہے۔ 15 دن میں تیار ہو جائے گا۔", "Urdu", "Prize", "Safe"),
    ("آپ کا پاسپورٹ تیار ہے۔ قریبی پاسپورٹ دفتر سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کا ڈرائیونگ لائسنس کی فیس موصول ہو گئی ہے۔ بائیو میٹرک کے لیے دفتر آئیں۔", "Urdu", "Prize", "Safe"),
    ("آپ کا ووٹرجسٹریشن کارڈ تیار ہے۔ قریبی دفتر سے شناختی کارڈ کے ساتھ حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی گاڑی کی رجسٹریشن مکمل ہو گئی ہے۔ نمبر پلیٹ 10 دن میں بھیج دی جائے گی۔", "Urdu", "Prize", "Safe"),
    ("آپ کا ٹوکن ٹیکس موصول ہو گیا ہے۔ رسید محفوظ رکھیں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی زمین کی فرد آن لائن دستیاب ہے۔ پنجاب لینڈ ریکارڈ اتھارٹی سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کے بچے کا ب فارم تیار ہے۔ قریبی دفتر سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی صحت انشورنس کی تجدید ہو گئی ہے۔ نئی کارڈ 7 دن میں بھیجا جائے گا۔", "Urdu", "Prize", "Safe"),
    ("آپ کی پینشن کی ماہانہ ادائیگی Rs. 25,000 اکاؤنٹ میں جمع ہو گئی ہے۔", "Urdu", "Prize", "Safe"),

    # ──────────────── ROMAN URDU TELECOM SAFE (anti-FP) ────────────────
    ("Zong: Aap ka SIM replacement complete ho gaya hai. Charges Rs. 200 deduct ho gaye hain. Naya SIM 2 ghante mein active hoga.", "Roman Urdu", "SIM Block", "Safe"),
    ("Jazz: Aap ka bundle kal expire ho raha hai. Renew karne ke liye *117# dial karein ya app se renew karein.", "Roman Urdu", "SIM Block", "Safe"),
    ("Zong: Sim replacement ke liye franchise visit karein. CNIC sath layein. Charges Rs. 200 apply honge.", "Roman Urdu", "SIM Block", "Safe"),
    ("Jazz: Aap ka monthly package kal expire ho jaye ga. Renew karne ke liye app use karein ya *117# dial karein.", "Roman Urdu", "SIM Block", "Safe"),
    ("Telenor: Aap ka package expire hone wala hai 2 din mein. Renew karne ke liye *555# dial karein.", "Roman Urdu", "SIM Block", "Safe"),
    ("Ufone: Aap ka balance Rs. 450 hai. Recharge karne ke liye *786# dial karein.", "Roman Urdu", "SIM Block", "Safe"),
    ("Zong: Aap ki call package successfully renew ho gayi hai. 3000 mins valid for 30 days.", "Roman Urdu", "SIM Block", "Safe"),
    ("Jazz: Aap ka 4G device package activate ho gaya hai. 10GB data 15 din ke liye.", "Roman Urdu", "SIM Block", "Safe"),
    ("Telenor: Aap ka postpaid bill Rs. 1,800 ready hai. Due date: 10 October. Pay via app.", "Roman Urdu", "SIM Block", "Safe"),
    ("Ufone: Aap ka super card auto-renew ho gaya hai. Rs. 1,200 deduct hue hain. 30 din valid.", "Roman Urdu", "SIM Block", "Safe"),

    # ──────────────── MORE PERSONAL/FAMILY SAFE (anti-FP) ────────────────
    ("Ammi: Kal nani aa rahi hain. Ghar saaf kar dena aur AC chala dena unke kamre mein.", "Roman Urdu", "Impersonation", "Safe"),
    ("Ammi: Beta kal subah jaldi uthna, humein hospital jaana hai Dadi ke checkup ke liye.", "Roman Urdu", "Impersonation", "Safe"),
    ("Ammi: Ghar mein AC chala dena, bohat garmi hai aaj. Main shaam tak aaungi.", "Roman Urdu", "Impersonation", "Safe"),
    ("Ammi: Kal ghar mein mehman aa rahe hain. Sab theek karke rakhna.", "Roman Urdu", "Impersonation", "Safe"),
    ("Ammi: Beta dhoodh le aana, fridge mein khatam ho gaya hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Ammi: Ghar aate waqt sabzi le aana. Aloo, tamatar aur pyaaz chahiye.", "Roman Urdu", "Impersonation", "Safe"),
    ("Beta ghar saaf kar dena, kal mehman aa rahe hain. Kamre mein AC bhi chala dena.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai, kal chutti hai. Chalein kahin ghumne? Sab log ready hain.", "Roman Urdu", "Impersonation", "Safe"),
    ("Papa: Office se aate waqt pharmacy se dawai le aana. Prescription table par hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Beti, kal shaam ko chai pe aana. Ghar mein sab khairiyat hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai, aaj ghar jaldi aana. Plumber aa raha hai bathroom fix karne.", "Roman Urdu", "Impersonation", "Safe"),
    ("Mom: Fridge mein khaana hai, garam karke kha lena. Main late aaungi aaj.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai, kal subah car ki servicing ke liye jaana hai. 8 bje ready rehna.", "Roman Urdu", "Impersonation", "Safe"),
    ("Ammi: Beta bahar se aate waqt bread le aana. Toast khatam ho gaya hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Papa: Kal chutti hai, ghar mein kuch repairs karwane hain. Plumber bulaya hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai, ghar ka AC kharab ho gaya hai. Kal technician aa raha hai.", "Roman Urdu", "Impersonation", "Safe"),
    ("Beta, ghar mein nani aa rahi hain kal. Unke kamre mein AC aur pani ready rakhna.", "Roman Urdu", "Impersonation", "Safe"),
    ("Sis: Kal chutti hai, chal shopping karte hain. Main 11 bje aaongi.", "Roman Urdu", "Impersonation", "Safe"),
    ("Bhai, kal subah 7 bje nikalna hai airport ke liye. Sab bags ready rakhna.", "Roman Urdu", "Impersonation", "Safe"),
    ("Papa: Ghar ki painting karwani hai next week. Colors decide kar lo.", "Roman Urdu", "Impersonation", "Safe"),

    # ──────────────── URDU SAFE: BANKING/FINANCIAL ────────────────
    ("آپ کے حبیب بینک اکاؤنٹ میں Rs. 25,000 جمع ہو گئے ہیں۔ موجودہ بیلنس Rs. 1,45,000۔ شکریہ۔", "Urdu", "Bank", "Safe"),
    ("آپ کا میزان بینک ڈیبٹ کارڈ تیار ہے۔ قریبی برانچ سے شناختی کارڈ کے ساتھ حاصل کریں۔", "Urdu", "Bank", "Safe"),
    ("ایچ بی ایل: آپ کی ماہانہ بینک سٹیٹمنٹ تیار ہے۔ آن لائن بینکنگ سے ڈاؤن لوڈ کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کے ایزی پیسہ اکاؤنٹ سے Rs. 3,500 بھیجے گئے ہیں۔ وصول کنندہ: احمد۔ ٹرانزیکشن آئی ڈی: EP7823456۔", "Urdu", "Bank", "Safe"),
    ("آپ کی ماہانہ تنخواہ Rs. 65,000 آپ کے اکاؤنٹ میں جمع ہو گئی ہے۔ رسید نمبر: SAL-2024-0892۔", "Urdu", "Bank", "Safe"),
    ("آپ کا جیز کیش اکاؤنٹ بیلنس Rs. 12,350 ہے۔ کیش آؤٹ کے لیے قریبی ایجنٹ سے رابطہ کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کے میزان بینک اکاؤنٹ کا منافع Rs. 4,200 جمع ہو گیا ہے۔ موجودہ بیلنس Rs. 2,15,000۔", "Urdu", "Bank", "Safe"),
    ("آپ کے بینک اکاؤنٹ کا سالانہ زکوۃ کٹوتی مکمل ہو گئی ہے۔ Rs. 1,200 زکوۃ جمع کروایا گیا۔", "Urdu", "Bank", "Safe"),
    ("آپ کی آن لائن شاپنگ کا آرڈر نمبر: DAR-89234۔ Rs. 4,500 داراز ایپ سے ٹریک کریں۔", "Urdu", "Bank", "Safe"),
    ("آپ کا کریڈٹ کارڈ کا بل Rs. 28,500 تیار ہے۔ آخری تاریخ 15 تاریخ۔ آن لائن ادا کریں۔", "Urdu", "Bank", "Safe"),
    ("ایف بی آر: آپ کا سالانہ ٹیکس ریٹرن کامیابی سے جمع ہو گیا ہے۔ رسید نمبر: FBR-2024-5678۔", "Urdu", "Bank", "Safe"),
    ("آپ کے نیا پے اکاؤنٹ کی تصدیق مکمل ہو گئی ہے۔ اب آپ تمام سروسز استعمال کر سکتے ہیں۔", "Urdu", "Bank", "Safe"),
    # ──────────────── URDU SAFE: TELECOM/UTILITIES ────────────────
    ("جاز: آپ کا ماہانہ پیکج Rs. 999 میں کامیابی سے تجدید ہو گیا ہے۔ 30 دن کے لیے 15 جی بی ڈیٹا فعال۔", "Urdu", "SIM Block", "Safe"),
    ("زونگ: آپ کا سم کارڈ کامیابی سے تبدیل ہو گیا ہے۔ Rs. 200 چارجز کٹے ہیں۔ نیا سم 2 گھنٹے میں فعال ہوگا۔", "Urdu", "SIM Block", "Safe"),
    ("ٹیلی نار: آپ کا پوسٹ پیڈ بل Rs. 2,100 ہے۔ آخری تاریخ 10 اکتوبر۔ ایزی پیسہ سے ادا کریں۔", "Urdu", "SIM Block", "Safe"),
    ("یوفون: آپ کا سپر کارڈ آٹو تجدید ہو گیا ہے۔ Rs. 1,200 کٹے ہیں۔ 30 دن تک فعال۔", "Urdu", "SIM Block", "Safe"),
    ("آپ کا بجلی کا بل Rs. 8,500 ہے۔ آخری تاریخ 25 تاریخ۔ لیسکو ایپ یا آن لائن ادا کریں۔", "Urdu", "SIM Block", "Safe"),
    ("آپ کا سوئی گیس کا بل Rs. 3,200 ہے۔ بینک یا ایپ کے ذریعے ادائیگی کریں۔", "Urdu", "SIM Block", "Safe"),
    ("آپ کا پی ٹی سی ایل انٹرنیٹ بل Rs. 4,500 ہے۔ 5 تاریخ تک ادا کریں۔", "Urdu", "SIM Block", "Safe"),
    ("آپ کا کے الیکٹرک بل Rs. 12,300 ہے۔ آخری تاریخ 20 تاریخ۔ آن لائن ادا کریں۔", "Urdu", "SIM Block", "Safe"),
    ("آپ کا واسا پانی کا بل Rs. 1,800 ہے۔ نیشنل بینک یا ایپ سے ادائیگی کریں۔", "Urdu", "SIM Block", "Safe"),
    ("آپ کا اسٹورم فائبر انٹرنیٹ بل Rs. 5,000 ہے۔ 15 تاریخ تک ادائیگی کریں۔", "Urdu", "SIM Block", "Safe"),
    # ──────────────── URDU SAFE: GOVERNMENT/NADRA ────────────────
    ("آپ کی نادرا شناختی کارڈ کی تجدید درخواست موصول ہو گئی ہے۔ 15 دن میں تیار ہو جائے گا۔", "Urdu", "Prize", "Safe"),
    ("آپ کا پاسپورٹ تیار ہے۔ قریبی پاسپورٹ دفتر سے شناختی کارڈ کے ساتھ حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کا ڈرائیونگ لائسنس کی فیس موصول ہو گئی ہے۔ بائیو میٹرک کے لیے دفتر آئیں۔", "Urdu", "Prize", "Safe"),
    ("بے نظیر انکم سپورٹ پروگرام: آپ کی اگلی قسط Rs. 12,000 موصول ہو گئی ہے۔ قریبی سینٹر سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کا صحت کارڈ فعال ہو گیا ہے۔ سرکاری ہسپتالوں میں مفت علاج کی سہولت دستیاب ہے۔", "Urdu", "Prize", "Safe"),
    ("احساس پروگرام: آپ کی درخواست منظور ہو گئی ہے۔ Rs. 14,000 ہر مہینے آپ کے اکاؤنٹ میں آئیں گے۔", "Urdu", "Prize", "Safe"),
    ("آپ کا ووٹر رجسٹریشن کارڈ تیار ہے۔ الیکشن کمیشن دفتر سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کے بچے کا ب فارم تیار ہے۔ قریبی نادرا دفتر سے شناختی کارڈ کے ساتھ حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی گاڑی کی رجسٹریشن مکمل ہو گئی ہے۔ نمبر پلیٹ 10 دن میں بھیج دی جائے گی۔", "Urdu", "Prize", "Safe"),
    ("آپ کا ٹوکن ٹیکس موصول ہو گیا ہے۔ رسید محفوظ رکھیں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی زمین کی فرد آن لائن دستیاب ہے۔ پنجاب لینڈ ریکارڈ اتھارٹی سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی پینشن کی ماہانہ ادائیگی Rs. 25,000 اکاؤنٹ میں جمع ہو گئی ہے۔", "Urdu", "Prize", "Safe"),
    ("آپ کی محکمہ صحت انشورنس کی تجدید ہو گئی ہے۔ نیا کارڈ 7 دن میں بھیجا جائے گا۔", "Urdu", "Prize", "Safe"),
    # ──────────────── URDU SAFE: PERSONAL/FAMILY ────────────────
    ("امی جان، میں گھر پہنچ گیا ہوں۔ پریشان نہ ہوں۔", "Urdu", "Impersonation", "Safe"),
    ("بیٹا، کل دادی کے گھر جانا ہے۔ تیار رہنا۔", "Urdu", "Impersonation", "Safe"),
    ("بھائی، کل عید کی خریداری کے لیے بازار چلیں گے۔ 10 بجے آ جانا۔", "Urdu", "Impersonation", "Safe"),
    ("بہن، کل بچوں کو اسکول چھوڑنا ہے۔ 7:30 بجے تیار رہنا۔", "Urdu", "Impersonation", "Safe"),
    ("ابو، آفس سے آتے ہوئے دودھ اور روٹی لے آنا۔", "Urdu", "Impersonation", "Safe"),
    ("امی، آج میری دوست آ رہی ہے۔ کھانا زیادہ بنایئں۔", "Urdu", "Impersonation", "Safe"),
    ("بھائی، کل میری شادی کی خریداری ہے۔ تم بھی چلنا۔", "Urdu", "Impersonation", "Safe"),
    ("بیٹا، گھر میں مہمان آئے ہوئے ہیں۔ جلدی آ جاؤ۔", "Urdu", "Impersonation", "Safe"),
    ("بہن، کل ہماری نانی کے گھر دعوت ہے۔ 12 بجے چلیں گی۔", "Urdu", "Impersonation", "Safe"),
    ("ابو، میری امتحان کی تیاری ہو گئی ہے۔ کل سے اسکول جاؤں گا۔", "Urdu", "Impersonation", "Safe"),
    ("امی جان، آج رات کھانا مت بنائیں۔ میں باہر سے بریانی لے آؤں گا۔", "Urdu", "Impersonation", "Safe"),
    ("بھائی جان، آج کرکٹ میچ ہے۔ شام کو 4 بجے پارک میں ملو۔", "Urdu", "Impersonation", "Safe"),
    ("بہن، بچوں کو لے کر کل ہمارے گھر آؤ۔ عید ملن پارٹی ہے۔", "Urdu", "Impersonation", "Safe"),
    ("بیٹا، دکان سے چینی اور چائے لے آؤ۔ فریج میں کھانا بھی ہے گرم کر کے کھا لینا۔", "Urdu", "Impersonation", "Safe"),
    ("امی، آج میری چھٹی ہے۔ کہیں گھومنے چلیں؟", "Urdu", "Impersonation", "Safe"),
    # ──────────────── URDU SAFE: HEALTH/EDUCATION ────────────────
    ("آپ کی ہسپتال اپائنٹمنٹ 15 تاریخ کو صبح 10 بجے ہے۔ شناختی کارڈ ساتھ لائیں۔", "Urdu", "Prize", "Safe"),
    ("آپ کے بچے کی ویکسین 20 تاریخ کو ہے۔ صحت مرکز پر شناختی کارڈ کے ساتھ آئیں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی لیبارٹری ٹیسٹ رپورٹ تیار ہے۔ ہسپتال کی ایپ سے ڈاؤن لوڈ کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کی ادویات کی آرڈر نمبر: PH-4523 تیار ہے۔ قریبی فارمیسی سے حاصل کریں۔", "Urdu", "Prize", "Safe"),
    ("آپ کے بچے کا اسکول فیس Rs. 15,000 موصول ہو گئی ہے۔ رسید نمبر: SCH-2024-0567۔", "Urdu", "Prize", "Safe"),
    ("آپ کی یونیورسٹی فیس Rs. 45,000 موصول ہو گئی ہے۔ سمسٹر رجسٹریشن مکمل۔", "Urdu", "Prize", "Safe"),
    ("آپ کا میڈیکل انشورنس کلیم منظور ہو گیا ہے۔ Rs. 85,000 ہسپتال کو ادا کیے گئے۔", "Urdu", "Prize", "Safe"),
    ("آپ کی ڈاکٹر اپائنٹمنٹ کی تصدیق ہو گئی ہے۔ 22 تاریخ 3 بجے کلینک پر آئیں۔", "Urdu", "Prize", "Safe"),
    # ──────────────── MIXED SAFE: BANKING/FINANCIAL ────────────────
    ("EasyPaisa Confirmation: Aap ka Rs. 5,000 ka transaction successful raha. Transaction ID: EP-45678901. Receiver: Ahmed Khan.", "Mixed", "Bank", "Safe"),
    ("JazzCash: Aap ke account mein Rs. 25,000 receive hue hain. Sender: Ali Enterprises. Receipt: JC-98765432.", "Mixed", "Bank", "Safe"),
    ("HBL Mobile: Aap ka bill payment Rs. 8,500 successful raha. Bill reference: HBL-2024-8901. Thank you for banking with us.", "Mixed", "Bank", "Safe"),
    ("Meezan Bank: Aap ka monthly statement ready hai. Internet banking ya app se download karein. Statement period: Sep 2024.", "Mixed", "Bank", "Safe"),
    ("Bank Alfalah: Aap ka debit card dispatch ke liye tayyar hai. Branch se ID ke saath collect karein. Card ending: 4523.", "Mixed", "Bank", "Safe"),
    ("UBL: Aap ki salary Rs. 75,000 credit ho gayi hai. Current balance: Rs. 2,45,000. Transaction ref: SAL-09234.", "Mixed", "Bank", "Safe"),
    ("SadaPay: Aap ka account successfully verify ho gaya hai. Ab aap saari services use kar sakte hain. Welcome to SadaPay!", "Mixed", "Bank", "Safe"),
    ("NayaPay: Aap ka Rs. 12,000 ka transfer successful raha. Receiver: Sara Ahmed. Ref: NP-34567890.", "Mixed", "Bank", "Safe"),
    ("FBR: Aap ka income tax return file ho gaya hai. Acknowledgement number: FBR-2024-78901. Status: Approved.", "Mixed", "Bank", "Safe"),
    ("Bank Al Habib: Aap ka fixed deposit mature ho gaya hai. Rs. 5,25,000 + profit Rs. 52,500. Branch visit karke renew karwayein.", "Mixed", "Bank", "Safe"),
    ("MCB: Aap ka loan installment Rs. 18,500 successfully deduct hua hai. Next due date: 5 November 2024.", "Mixed", "Bank", "Safe"),
    ("Askari Bank: Aap ka cheque number 45234567 clear ho gaya hai. Rs. 85,000 credit hue hain.", "Mixed", "Bank", "Safe"),
    # ──────────────── MIXED SAFE: TELECOM/UTILITIES ────────────────
    ("Jazz: Aap ka monthly bundle Rs. 1,199 mein successfully renew ho gaya hai. 20GB data + 3000 mins active hain. Valid: 30 din.", "Mixed", "SIM Block", "Safe"),
    ("Zong: Aap ka SIM replacement complete ho gaya hai. Charges Rs. 200 deduct ho gaye hain. Naya SIM 2 ghante mein active hoga.", "Mixed", "SIM Block", "Safe"),
    ("Telenor: Aap ka postpaid bill Rs. 2,350 ready hai. Due date: 15 October. JazzCash ya app se pay karein.", "Mixed", "SIM Block", "Safe"),
    ("Ufone: Aap ka super card Rs. 1,200 mein auto-renew ho gaya hai. 5GB data + unlimited calls. 30 din valid.", "Mixed", "SIM Block", "Safe"),
    ("LESCO: Aap ka electricity bill Rs. 9,800 generate ho gaya hai. Due date: 25 October. Online ya bank se pay karein.", "Mixed", "SIM Block", "Safe"),
    ("SNGPL: Aap ka gas bill Rs. 4,200 hai. Last date: 20 October. JazzCash ya EasyPaisa se pay karein.", "Mixed", "SIM Block", "Safe"),
    ("PTCL: Aap ka internet + landline bill Rs. 5,500 ready hai. Due date: 18 October. PTCL app se pay karein.", "Mixed", "SIM Block", "Safe"),
    ("K-Electric: Aap ka bill Rs. 14,500 generate hua hai. Due date: 22 October. Online ya bank se pay karein.", "Mixed", "SIM Block", "Safe"),
    ("StormFiber: Aap ka monthly bill Rs. 4,500 ready hai. Due date: 10 November. App ya bank se pay karein.", "Mixed", "SIM Block", "Safe"),
    ("WASA: Aap ka water bill Rs. 2,100 generate ho gaya hai. Due date: 30 October. Bank ya app se pay karein.", "Mixed", "SIM Block", "Safe"),
    # ──────────────── MIXED SAFE: PERSONAL/FAMILY ────────────────
    ("Bhai, ghar aate waqt pharmacy se Dolo 650 le aana. Papa ko bukhar hai. Prescription table par hai.", "Mixed", "Impersonation", "Safe"),
    ("Ammi: Beta kal subah 7 bje nikalna hai, humein Nani ke ghar jaana hai. Nashta karke aana.", "Mixed", "Impersonation", "Safe"),
    ("Bhai, kal Sunday hai. Chalo cricket match dekhte hain National Stadium mein. Tickets ready hain.", "Mixed", "Impersonation", "Safe"),
    ("Beti, kal ghar aa jana. Tumhare pasand ki biryani banaongi. Papa bhi tumhara intezar kar rahe hain.", "Mixed", "Impersonation", "Safe"),
    ("Bhai, ghar ka pani ka pump kharab ho gaya hai. Plumber ko call kar diya hai, shaam tak aa jaye ga.", "Mixed", "Impersonation", "Safe"),
    ("Ammi: Beta fridge mein khaana hai, garam karke kha lena. Main shaam 7 bje tak ghar aaungi.", "Mixed", "Impersonation", "Safe"),
    ("Bhai, kal chutti hai. Chalo Murree ghumne chalte hain? Sab log ready hain, subah 6 bje nikalte hain.", "Mixed", "Impersonation", "Safe"),
    ("Mom: Ghar mein AC theek ho gaya hai. Kal se use kar sakte ho. Technician ne check kar liya.", "Mixed", "Impersonation", "Safe"),
    ("Papa: Beta, kal subah jaldi uthna, humein hospital jaana hai Dadi ke checkup ke liye. 8 bje ready rehna.", "Mixed", "Impersonation", "Safe"),
    ("Sis: Kal meri friend ki wedding hai. Chal shopping karte hain aaj shaam. 5 bje Liberty Market mein milte hain.", "Mixed", "Impersonation", "Safe"),
    ("Bhai, ghar ki painting ka kaam shuru ho gaya hai. Kal tak living room complete ho jaye ga.", "Mixed", "Impersonation", "Safe"),
    ("Beta, ghar mein nani aa rahi hain kal. Unke kamre mein AC aur pani ready rakhna. Unki dawai bhi rakh dena.", "Mixed", "Impersonation", "Safe"),
    ("Ammi: Kal chutti hai, ghar mein kuch repairs karwane hain. Plumber aur electrician bulaye hain.", "Mixed", "Impersonation", "Safe"),
    ("Bhai, car ki servicing kal karwani hai. 9 bje Honda service center mein jaana hai. Ready rehna.", "Mixed", "Impersonation", "Safe"),
    ("Papa: Beta, office se aate waqt grocery le aana. Doodh, bread, anday aur sabzi chahiye.", "Mixed", "Impersonation", "Safe"),
    # ──────────────── MIXED SAFE: GOVERNMENT/NADRA ────────────────
    ("NADRA: Aap ka CNIC renew ho gaya hai. Collection ke liye qareebi NADRA office visit karein. CNIC ready hai.", "Mixed", "Prize", "Safe"),
    ("BISP: Aap ki quarterly payment Rs. 14,000 aap ke account mein credit ho gayi hai. JazzCash se check karein.", "Mixed", "Prize", "Safe"),
    ("Ehsaas Program: Aap ki application approve ho gayi hai. Monthly stipend Rs. 12,000 start ho jaye ga.", "Mixed", "Prize", "Safe"),
    ("PTA: Aap ka device successfully register ho gaya hai. Registration ID: PTA-2024-567890. Ab network use kar sakte hain.", "Mixed", "Prize", "Safe"),
    ("SEHAT Card: Aap ka health insurance card active ho gaya hai. Government hospitals mein free treatment available hai.", "Mixed", "Prize", "Safe"),
    ("FBR: Aap ka NTN successfully register ho gaya hai. NTN: 1234567-8. Tax filing ab online kar sakte hain.", "Mixed", "Prize", "Safe"),
    ("NADRA: Aap ke bachay ka B-Form tayyar hai. Qareebi NADRA office se ID ke saath collect karein.", "Mixed", "Prize", "Safe"),
    ("Punjab Government: Aap ka rasta ban gaya hai. Gali number 5 ki repair ka kaam shuru ho gaya hai.", "Mixed", "Prize", "Safe"),
    ("Driving License: Aap ka license renew ho gaya hai. Collection ke liye traffic police office visit karein.", "Mixed", "Prize", "Safe"),
    ("Passport Office: Aap ka passport ready hai. Qareebi passport office se CNIC ke saath collect karein.", "Mixed", "Prize", "Safe"),
    # ──────────────── MIXED SAFE: SERVICE/HEALTH ────────────────
    ("Careem: Aap ki ride complete ho gayi hai. Fare: Rs. 850. Payment: Cash. Rating dein app se. Shukriya!", "Mixed", "Prize", "Safe"),
    ("FoodPanda: Aap ka order #FP-45678 deliver ho gaya hai. Total: Rs. 1,250. Enjoy your meal!", "Mixed", "Prize", "Safe"),
    ("Daraz.pk: Aap ka order #DAR-89012 ship ho gaya hai. Expected delivery: 3-5 working days. Track via app.", "Mixed", "Prize", "Safe"),
    ("InDrive: Aap ki ride complete ho gayi hai. Fare: Rs. 650 paid via cash. Driver ko rate karein.", "Mixed", "Prize", "Safe"),
    ("Bykea: Aap ki parcel delivery complete ho gayi hai. Charges: Rs. 350. Receipt app mein check karein.", "Mixed", "Prize", "Safe"),
    ("Hospital: Aap ki appointment 15 October ko 10 AM hai. Dr. Ahmed Khan. ID card sath layein.", "Mixed", "Prize", "Safe"),
    ("Lab Results: Aap ki blood test report ready hai. Hospital app se download karein. Reference: LAB-2024-5678.", "Mixed", "Prize", "Safe"),
    ("School: Aap ke bachay ki fees Rs. 18,000 receive ho gayi hain. Receipt: SCH-2024-0912. Shukriya.", "Mixed", "Prize", "Safe"),
    ("University: Aap ki semester fees Rs. 55,000 receive ho gayi hai. Registration complete. Ref: UNI-2024-3456.", "Mixed", "Prize", "Safe"),
    ("TCS: Aap ka parcel successfully deliver ho gaya hai. Tracking ID: TCS-78901234. Thank you for using TCS.", "Mixed", "Prize", "Safe"),
    ("Leopards Courier: Aap ka shipment tracking ID: LEO-567890. Status: Out for delivery. Expected today.", "Mixed", "Prize", "Safe"),
    ("Pakistan Post: Aap ka registered parcel receive ho gaya hai. Tracking: PKG-345678. Post office se collect karein.", "Mixed", "Prize", "Safe"),
]

# ── build workbook ──────────────────────────────────────────────────────────
wb = Workbook()
ws = wb.active
ws.title = "Scam Detection Dataset"

# Headers
headers = ["Message Content", "Language Type", "Scam Category", "Label"]
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_idx, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Combine & shuffle
all_data = SCAM_MESSAGES + SAFE_MESSAGES
random.seed(42)
random.shuffle(all_data)

# Write rows
wrap_align = Alignment(vertical="top", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="top")

scam_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
safe_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")

for row_idx, (msg, lang, cat, label) in enumerate(all_data, 2):
    ws.cell(row=row_idx, column=1, value=msg).alignment = wrap_align
    ws.cell(row=row_idx, column=2, value=lang).alignment = center_align
    ws.cell(row=row_idx, column=3, value=cat).alignment = center_align
    ws.cell(row=row_idx, column=4, value=label).alignment = center_align

    fill = scam_fill if label == "Scam" else safe_fill
    for col in range(1, 5):
        ws.cell(row=row_idx, column=col).fill = fill
        ws.cell(row=row_idx, column=col).border = thin_border

# Deduplicate (by message text, case-insensitive)
seen = set()
rows_to_delete = []
for row_idx in range(2, ws.max_row + 1):
    msg = ws.cell(row=row_idx, column=1).value.strip().lower()
    h = hashlib.md5(msg.encode()).hexdigest()
    if h in seen:
        rows_to_delete.append(row_idx)
    else:
        seen.add(h)

for idx in sorted(rows_to_delete, reverse=True):
    ws.delete_rows(idx)

# Column widths
ws.column_dimensions["A"].width = 80
ws.column_dimensions["B"].width = 16
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 10

# Freeze top row
ws.freeze_panes = "A2"

# Auto-filter
ws.auto_filter.ref = f"A1:D{ws.max_row}"

# ── Stats sheet ──────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Dataset Statistics")
ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True)
ws2.cell(row=1, column=2, value="Count").font = Font(bold=True)

total_rows = ws.max_row - 1  # minus header
scam_count = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(row=r, column=4).value == "Scam")
safe_count = total_rows - scam_count

# Language distribution
lang_counts = {}
cat_counts = {}
for r in range(2, ws.max_row + 1):
    lang = ws.cell(row=r, column=2).value
    cat = ws.cell(row=r, column=3).value
    lang_counts[lang] = lang_counts.get(lang, 0) + 1
    cat_counts[cat] = cat_counts.get(cat, 0) + 1

stats = [
    ("Total Messages", total_rows),
    ("Scam Messages", scam_count),
    ("Safe Messages", safe_count),
    ("", ""),
    ("── Language Distribution ──", ""),
]
for lang in sorted(lang_counts):
    stats.append((f"  {lang}", lang_counts[lang]))
stats.append(("", ""))
stats.append(("── Category Distribution ──", ""))
for cat in sorted(cat_counts):
    stats.append((f"  {cat}", cat_counts[cat]))

for i, (metric, count) in enumerate(stats, 2):
    ws2.cell(row=i, column=1, value=metric)
    ws2.cell(row=i, column=2, value=count)

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 12

# Save
OUTPUT = r"d:\Ai Hackaton\scam_detection\data\scam_messages_dataset.xlsx"
COPY   = r"d:\Ai Hackaton\scam_messages_dataset.xlsx"
import shutil
wb.save(OUTPUT)
shutil.copy2(OUTPUT, COPY)
print(f"Dataset saved to: {OUTPUT}")
print(f"Copy saved to: {COPY}")
print(f"Total messages: {total_rows}")
print(f"  Scam: {scam_count}")
print(f"  Safe: {safe_count}")
print(f"Languages: {dict(sorted(lang_counts.items()))}")
print(f"Categories: {dict(sorted(cat_counts.items()))}")
print(f"Generated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
